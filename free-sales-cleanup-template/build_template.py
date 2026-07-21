# Build "Sales Cleanup Template - Free (Automated Workflow LLC)" as xlsx for Drive conversion
# Polished to AWLLC brand standards: banner bands, color-coded KPIs, card fills, wordmark footers.
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ---- brand tokens (single source of truth) ----
import sys as _sys, os as _os
_sys.path.insert(0, _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "..", "_brand"))
from awllc_brand import INK, GREEN, AMBER, RED, MUTE, CARD, BANDBG, WORDMARK, GREENBG, REDBG, AMBERBG  # single source of truth
CREAM = CARD  # legacy alias (== WELL)

HEADER_FILL = PatternFill("solid", fgColor=INK)
CARD_FILL   = PatternFill("solid", fgColor=CARD)
GREEN_FILL  = PatternFill("solid", fgColor=GREENBG)
BAND_FILL   = PatternFill("solid", fgColor=BANDBG)
thin = Side(style="thin", color="E4DFD1")
border = Border(bottom=thin)

def band(ws, rng, text, size=14):
    """Dark banner across a merged range with white title text."""
    ws.merge_cells(rng)
    top = rng.split(":")[0]
    ws[top] = text
    ws[top].font = Font(bold=True, size=size, color="FFFFFF")
    ws[top].alignment = Alignment(horizontal="left", vertical="center")
    # fill every cell in the merged range so the band is solid
    c1 = rng.split(":")[0]; c2 = rng.split(":")[1]
    from openpyxl.utils.cell import range_boundaries
    minc, minr, maxc, maxr = range_boundaries(rng)
    for r in range(minr, maxr+1):
        for c in range(minc, maxc+1):
            ws.cell(r, c).fill = HEADER_FILL
    ws.row_dimensions[minr].height = 30

def wordmark(ws, row, span_last_col="B"):
    ws.cell(row, 1, "Built by Automated Workflow LLC  ·  automatedworkflowllc.com  ·  free template")
    ws.cell(row, 1).font = Font(italic=True, size=9, color=WORDMARK)

# ---------- Sales data ----------
ws = wb.active
ws.title = "Sales data"

headers = ["Date", "Invoice #", "Client", "Category", "Amount", "Status"]
rows = [
    # April  (sum 18,400)
    (dt.date(2026,4,2),  "INV-1001", "Hogtown Coffee Roasters",   "Install Kits",     2400, "Paid"),
    (dt.date(2026,4,6),  "INV-1002", "Paynes Prairie Landscaping","Consulting",       1800, "Paid"),
    (dt.date(2026,4,9),  "INV-1003", "Depot Ave Bike Shop",       "Maintenance Plan",  950, "Paid"),
    (dt.date(2026,4,13), "INV-1004", "Micanopy Antiques Co.",     "Install Kits",     3200, "Paid"),
    (dt.date(2026,4,16), "INV-1005", "Sweetwater Dental Studio",  "Training",         1200, "Paid"),
    (dt.date(2026,4,20), "INV-1006", "Archer Road Fitness",       "consulting",       1800, "Paid"),      # miscat: lowercase
    (dt.date(2026,4,23), "INV-1007", "Newnans Lake Outfitters",   "Install Kits",     2750, "Paid"),
    (dt.date(2026,4,27), "INV-1003", "Depot Ave Bike Shop",       "Maintenance Plan",  950, "Paid"),      # DUP INV-1003
    (dt.date(2026,4,29), "INV-1008", "Tioga Town Bakery",         "Maintenance Plan", 3350, "Paid"),
    # May (sum 21,750)
    (dt.date(2026,5,4),  "INV-1009", "Haile Village Realty",      "Install Kits",     2900, "Paid"),
    (dt.date(2026,5,7),  "INV-1010", "Kanapaha Gardens Nursery",  "Consulting ",      1500, "Paid"),      # miscat: trailing space
    (dt.date(2026,5,11), "INV-1011", "Duckpond Design Studio",    "Training",         1100, "Paid"),
    (dt.date(2026,5,14), "INV-1012", "Alachua Auto Care",         "Install Kits",     3400, "Paid"),
    (dt.date(2026,5,18), "INV-1013", "Millhopper Vet Supply",     "Maintenance Plan", 1250, "Overdue"),
    (dt.date(2026,5,20), "INV-1012", "Alachua Auto Care",         "Install Kits",     3400, "Paid"),      # DUP INV-1012
    (dt.date(2026,5,22), "INV-1014", "Santa Fe Print Shop",       "Consulting",       2100, "Paid"),
    (dt.date(2026,5,26), "INV-1015", "Cedar Key Seafood Market",  "Install Kits",     2800, "Paid"),
    (dt.date(2026,5,28), "INV-1016", "University Ave Books",      "Maintenance Plan", 3300, "Paid"),
    # June (sum 24,900)
    (dt.date(2026,6,1),  "INV-1017", "Hawthorne Trail Cycles",    "Install Kits",     3100, "Paid"),
    (dt.date(2026,6,4),  "INV-1018", "Prairie Creek Builders",    "Consutling",        900, "Paid"),      # miscat: typo
    (dt.date(2026,6,8),  "INV-1019", "Ridgeway Cleaning Co.",     "Maintenance Plan", 1400, "Unpaid"),
    (dt.date(2026,6,10), "INV-1020", "Loblolly Landscaping",      "Consulting",       3400, "Paid"),
    (dt.date(2026,6,15), "INV-1021", "Westside Yoga Collective",  "Training",         1300, "Paid"),
    (dt.date(2026,6,17), "INV-1022", "Buckman Hardware",          "Install Kits",     3600, "Paid"),
    (dt.date(2026,6,19), "INV-1021", "Westside Yoga Collective",  "Training",         1300, "Paid"),      # DUP INV-1021
    (dt.date(2026,6,23), "INV-1023", "Rocky Point Storage",       "Install Kits",     3950, "Paid"),
    (dt.date(2026,6,26), "INV-1024", "Glen Springs Media",        "Install Kits",     5950, "Paid"),
    # Blank dates (sum 1,900)
    (None,               "INV-1025", "Two Tails Pet Grooming",    "Training",          850, "Paid"),
    (None,               "INV-1026", "Osprey Point Cafe",         "Maintenance Plan",  600, "Paid"),
    (None,               "INV-1027", "High Springs Hardware",     "Consulting",        450, "Overdue"),
]

for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = HEADER_FILL
ws.row_dimensions[1].height = 24

STATUS_COL = 6
for r, row in enumerate(rows, 2):
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = border
        if c == 1 and v is not None:
            cell.number_format = "m/d/yyyy"
        if c == 5:
            cell.number_format = "$#,##0"
    # tint the status cell for non-paid rows so problems read at a glance
    st = row[5]
    if st in ("Unpaid", "Overdue"):
        sc = ws.cell(row=r, column=STATUS_COL)
        sc.font = Font(bold=True, color=RED)

ws.freeze_panes = "A2"
widths = [12, 12, 28, 18, 12, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
wordmark(ws, 33)

SD = "'Sales data'"
DATA_E = f"{SD}!$E$2:$E$31"
DATA_A = f"{SD}!$A$2:$A$31"
DATA_B = f"{SD}!$B$2:$B$31"
DATA_D = f"{SD}!$D$2:$D$31"
DATA_F = f"{SD}!$F$2:$F$31"
NOT_CANON = (f"(1-(EXACT({DATA_D},\"Consulting\")+EXACT({DATA_D},\"Install Kits\")"
             f"+EXACT({DATA_D},\"Maintenance Plan\")+EXACT({DATA_D},\"Training\")))")

# ---------- Dashboard ----------
db = wb.create_sheet("Dashboard")
band(db, "A1:B1", "SALES DASHBOARD")
db["A2"] = "Auto-updates from the Sales data tab. Add rows there and every number below recalculates."
db["A2"].font = Font(italic=True, size=10, color=MUTE)

# KPI card
kpis = [
    ("Revenue (Apr-Jun)", "=SUM(B11:B13)", INK),
    ("Collected (Paid)", f"=SUMIF({DATA_F},\"Paid\",{DATA_E})", GREEN),
    ("Outstanding (Unpaid + Overdue)", f"=SUMIF({DATA_F},\"Unpaid\",{DATA_E})+SUMIF({DATA_F},\"Overdue\",{DATA_E})", RED),
    ("Recoverable found by Data Health", "='Data Health'!C4", GREEN),
]
for i, (label, f, color) in enumerate(kpis):
    r = 4 + i
    lc = db.cell(row=r, column=1, value=label)
    lc.font = Font(bold=True, color=INK)
    lc.fill = CARD_FILL
    c = db.cell(row=r, column=2, value=f)
    c.number_format = "$#,##0"
    c.font = Font(bold=True, size=13, color=color)
    c.fill = CARD_FILL
    c.alignment = Alignment(horizontal="right")

def section(cell_addr, text):
    db[cell_addr] = text
    db[cell_addr].font = Font(bold=True, color=INK)
    db[cell_addr].fill = BAND_FILL
    # band the B column too
    col = cell_addr[0]; row = cell_addr[1:]
    db[f"B{row}"].fill = BAND_FILL

section("A10", "Revenue by month"); db["B10"] = "Revenue"
db["B10"].font = Font(bold=True, color=INK); db["B10"].fill = BAND_FILL
db["B10"].alignment = Alignment(horizontal="right")
months = [("April", 4, 5), ("May", 5, 6), ("June", 6, 7)]
for i, (name, m1, m2) in enumerate(months):
    r = 11 + i
    db.cell(row=r, column=1, value=name)
    f = (f"=SUMIFS({DATA_E},{DATA_A},\">=\"&DATE(2026,{m1},1),{DATA_A},\"<\"&DATE(2026,{m2},1))")
    c = db.cell(row=r, column=2, value=f)
    c.number_format = "$#,##0"; c.alignment = Alignment(horizontal="right")
db["A14"] = "Rows with a blank Date never land in a month - that's how sales drop out of the quarter (see Data Health)."
db["A14"].font = Font(italic=True, size=9, color=MUTE)

section("A16", "Revenue by category"); db["B16"] = "Revenue"
db["B16"].font = Font(bold=True, color=INK); db["B16"].fill = BAND_FILL
db["B16"].alignment = Alignment(horizontal="right")
cats = ["Consulting", "Install Kits", "Maintenance Plan", "Training"]
for i, cat in enumerate(cats):
    r = 17 + i
    db.cell(row=r, column=1, value=cat)
    c = db.cell(row=r, column=2, value=f"=SUMIF({DATA_D},A{r},{DATA_E})")
    c.number_format = "$#,##0"; c.alignment = Alignment(horizontal="right")
db["A21"] = "Not matching any category above"
db["A21"].font = Font(bold=True, color=AMBER)
c = db["B21"]; c.value = f"=SUM({DATA_E})-SUM(B17:B20)"
c.number_format = "$#,##0"; c.font = Font(bold=True, color=AMBER); c.alignment = Alignment(horizontal="right")
db["A22"] = "Money hiding under a mistyped label -> the Data Health tab lists how much."
db["A22"].font = Font(italic=True, size=9, color=MUTE)
wordmark(db, 24)

db.column_dimensions["A"].width = 38
db.column_dimensions["B"].width = 16

# ---------- Data Health ----------
dh = wb.create_sheet("Data Health")
band(dh, "A1:C1", "DATA HEALTH")
dh["A2"] = "Automatic checks on the Sales data tab. Fix the flagged rows and this money comes back into your totals."
dh["A2"].font = Font(italic=True, size=10, color=MUTE)

# headline "money found" panel
dh.merge_cells("A4:B4")
dh["A4"] = "RECOVERABLE REVENUE FOUND"
dh["A4"].font = Font(bold=True, size=13, color=GREEN)
dh["A4"].alignment = Alignment(horizontal="left", vertical="center")
c = dh["C4"]; c.value = "=C7+C8"; c.number_format = "$#,##0"
c.font = Font(bold=True, size=16, color=GREEN); c.alignment = Alignment(horizontal="right", vertical="center")
for addr in ("A4", "B4", "C4"):
    dh[addr].fill = GREEN_FILL
dh.row_dimensions[4].height = 30

for col, h in zip("ABC", ["Check", "Flagged", "Dollars affected"]):
    cell = dh[f"{col}6"]
    cell.value = h
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = HEADER_FILL

# (label, count formula, dollar formula or None, dollar color)
checks = [
    ("Miscategorized revenue (label doesn't match the master category list)",
     f"=SUMPRODUCT({NOT_CANON})",
     f"=SUMPRODUCT({DATA_E},{NOT_CANON})", AMBER),
    ("Blank sale dates (rows dropping out of the monthly totals)",
     f"=COUNTBLANK({DATA_A})",
     f"=SUMIFS({DATA_E},{DATA_A},\"\")", AMBER),
    ("Duplicate invoice numbers (same sale counted twice)",
     f"=SUMPRODUCT((COUNTIF({DATA_B},{DATA_B})>1)/COUNTIF({DATA_B},{DATA_B}))",
     None, AMBER),
    ("Unpaid / overdue invoices (billed but not collected)",
     f"=COUNTIF({DATA_F},\"Unpaid\")+COUNTIF({DATA_F},\"Overdue\")",
     f"=SUMIF({DATA_F},\"Unpaid\",{DATA_E})+SUMIF({DATA_F},\"Overdue\",{DATA_E})", RED),
]
for i, (label, fcount, fdollar, dcolor) in enumerate(checks):
    r = 7 + i
    zebra = PatternFill("solid", fgColor="F4F1E8") if i % 2 == 0 else None
    lc = dh.cell(row=r, column=1, value=label)
    cc = dh.cell(row=r, column=2, value=fcount)
    cc.alignment = Alignment(horizontal="center")
    cc.font = Font(bold=True, color=INK)
    if fdollar:
        c = dh.cell(row=r, column=3, value=fdollar)
        c.number_format = "$#,##0"; c.font = Font(bold=True, color=dcolor)
        c.alignment = Alignment(horizontal="right")
    else:
        dc = dh.cell(row=r, column=3, value="review & delete repeats")
        dc.font = Font(italic=True, color=MUTE, size=9)
    for col in range(1, 4):
        cell = dh.cell(row=r, column=col)
        cell.border = border
        if zebra:
            cell.fill = zebra

dh["A12"] = "Recoverable = miscategorized dollars + blank-date dollars (rows 7-8). Duplicates are a row count - delete the repeats so nothing is double-counted."
dh["A12"].font = Font(italic=True, size=9, color=MUTE)
dh["A13"] = "Unpaid/overdue is tracked separately - that's collections follow-up, not miscounting."
dh["A13"].font = Font(italic=True, size=9, color=MUTE)
dh["A15"] = "Master categories: Consulting, Install Kits, Maintenance Plan, Training. Renamed your categories? Update them inside the two formulas in row 7."
dh["A15"].font = Font(italic=True, size=9, color=MUTE)
wordmark(dh, 17)

dh.column_dimensions["A"].width = 62
dh.column_dimensions["B"].width = 10
dh.column_dimensions["C"].width = 18

path = r"C:\Users\hisbo\Documents\awllc-website\free-sales-cleanup-template\sales-cleanup-template.xlsx"
wb.save(path)

# sanity math
amts = [r[4] for r in rows]
print("rows:", len(rows), "total:", sum(amts))
print("apr/may/jun:", sum(a for r, a in zip(rows, amts) if r[0] and r[0].month == 4),
      sum(a for r, a in zip(rows, amts) if r[0] and r[0].month == 5),
      sum(a for r, a in zip(rows, amts) if r[0] and r[0].month == 6))
canon = {"Consulting", "Install Kits", "Maintenance Plan", "Training"}
print("miscat $:", sum(a for r, a in zip(rows, amts) if r[3] not in canon))
print("blankdate $:", sum(a for r, a in zip(rows, amts) if r[0] is None))
print("recoverable $:", sum(a for r, a in zip(rows, amts) if r[3] not in canon) + sum(a for r, a in zip(rows, amts) if r[0] is None))
print("unpaid/overdue $:", sum(a for r, a in zip(rows, amts) if r[5] != "Paid"))
print("saved:", path)
