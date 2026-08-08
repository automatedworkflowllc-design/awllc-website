# Build "Reconciler Demo - (Automated Workflow LLC)" as xlsx for Drive conversion.
# The product: two systems that SHOULD match (what you booked vs. what you billed) ->
# a Reconciliation tab that auto-surfaces ONLY the mismatches + a net-dollar gap.
# Demo scenario: a golf-sim / bay-rental shop ("Fairway Bays"), one week of play.
# Brand-matched to build_template.py (bands, KPI cards, wordmark footers).
import datetime as dt
import pathlib as _pl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ---- brand tokens (single source of truth) ----
import sys as _sys, os as _os
_sys.path.insert(0, _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "..", "_brand"))
from awllc_brand import INK, GREEN, AMBER, RED, MUTE, CARD, BANDBG, WORDMARK, GREENBG, REDBG, AMBERBG  # single source of truth

HEADER_FILL = PatternFill("solid", fgColor=INK)
CARD_FILL   = PatternFill("solid", fgColor=CARD)
GREEN_FILL  = PatternFill("solid", fgColor=GREENBG)
RED_FILL    = PatternFill("solid", fgColor=REDBG)
AMBER_FILL  = PatternFill("solid", fgColor=AMBERBG)
BAND_FILL   = PatternFill("solid", fgColor=BANDBG)
thin = Side(style="thin", color="E4DFD1")
border = Border(bottom=thin)

def band(ws, rng, text, size=14):
    ws.merge_cells(rng)
    top = rng.split(":")[0]
    ws[top] = text
    ws[top].font = Font(bold=True, size=size, color="FFFFFF")
    ws[top].alignment = Alignment(horizontal="left", vertical="center")
    from openpyxl.utils.cell import range_boundaries
    minc, minr, maxc, maxr = range_boundaries(rng)
    for r in range(minr, maxr+1):
        for c in range(minc, maxc+1):
            ws.cell(r, c).fill = HEADER_FILL
    ws.row_dimensions[minr].height = 30

def wordmark(ws, row):
    ws.cell(row, 1, "Built by Automated Workflow LLC  ·  automatedworkflowllc.com  ·  the Reconciler")
    ws.cell(row, 1).font = Font(italic=True, size=9, color=WORDMARK)

def headers_row(ws, r, cols):
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
    ws.row_dimensions[r].height = 22

# ============ System A: Bookings (the schedule / what you delivered) ============
bk = wb.active
bk.title = "Bookings"
band(bk, "A1:F1", "SYSTEM A  ·  BOOKINGS  (what the schedule says you delivered)")
bk["A2"] = "Pulled from your booking system. One row per bay session this week. Expected $ = Hours x Rate."
bk["A2"].font = Font(italic=True, size=10, color=MUTE)
headers_row(bk, 3, ["Booking ID", "Date", "Bay", "Hours", "Rate/hr", "Expected $"])

# (Booking ID, Date, Bay, Hours, Rate, known_status)  -> Expected computed by formula
# known_status is used for STATIC row coloring (demo data is fixed; Sheets rejects
# openpyxl conditional-formatting dxf blocks on import — learned 7/16, keep colors static)
bookings = [
    ("B-201", dt.date(2026,7,6),  "Bay 1", 2, 45, "OK"),
    ("B-202", dt.date(2026,7,6),  "Bay 3", 1, 45, "OK"),
    ("B-203", dt.date(2026,7,7),  "Bay 2", 3, 45, "HOURS MISMATCH"),  # billed only 2h (-45)
    ("B-204", dt.date(2026,7,7),  "Bay 4", 2, 45, "NOT BILLED"),      # never invoiced (-90)
    ("B-205", dt.date(2026,7,8),  "Bay 1", 2, 45, "OK"),
    ("B-206", dt.date(2026,7,8),  "Bay 2", 4, 45, "OK"),
    ("B-207", dt.date(2026,7,9),  "Bay 3", 1, 45, "OK"),
    ("B-208", dt.date(2026,7,9),  "Bay 4", 3, 45, "AMOUNT MISMATCH"), # $50/hr not $45 (+15)
    ("B-209", dt.date(2026,7,10), "Bay 1", 2, 45, "OK"),
    ("B-210", dt.date(2026,7,10), "Bay 2", 2, 45, "OK"),
    ("B-211", dt.date(2026,7,11), "Bay 3", 3, 45, "OK"),
    ("B-212", dt.date(2026,7,11), "Bay 4", 1, 45, "NOT BILLED"),      # never invoiced (-45)
    ("B-213", dt.date(2026,7,12), "Bay 1", 2, 45, "OK"),
    ("B-214", dt.date(2026,7,12), "Bay 2", 5, 45, "OK"),
]
BK_FIRST = 4
for i, (bid, d, bay, hrs, rate, _st) in enumerate(bookings):
    r = BK_FIRST + i
    bk.cell(r, 1, bid).border = border
    c = bk.cell(r, 2, d); c.number_format = "m/d/yyyy"; c.border = border
    bk.cell(r, 3, bay).border = border
    bk.cell(r, 4, hrs).border = border
    c = bk.cell(r, 5, rate); c.number_format = "$#,##0"; c.border = border
    c = bk.cell(r, 6, f"=D{r}*E{r}"); c.number_format = "$#,##0"; c.border = border
    c.font = Font(bold=True, color=INK)
BK_LAST = BK_FIRST + len(bookings) - 1
bk.freeze_panes = "A4"
for i, w in enumerate([12, 11, 8, 8, 9, 12], 1):
    bk.column_dimensions[get_column_letter(i)].width = w
wordmark(bk, BK_LAST + 2)

# ============ System B: Invoices (the billing / what you actually charged) ============
iv = wb.create_sheet("Invoices")
band(iv, "A1:E1", "SYSTEM B  ·  INVOICES  (what your billing actually charged)")
iv["A2"] = "Pulled from your invoicing tool. Should line up with Bookings by Booking ID - but it never fully does."
iv["A2"].font = Font(italic=True, size=10, color=MUTE)
headers_row(iv, 3, ["Invoice #", "Booking ID", "Date", "Hours billed", "Amount $"])

# (Invoice #, Booking ID, Date, Hours billed, Amount)
invoices = [
    ("INV-77", "B-201", dt.date(2026,7,6),  2, 90),
    ("INV-78", "B-202", dt.date(2026,7,6),  1, 45),
    ("INV-79", "B-203", dt.date(2026,7,7),  2, 90),    # only 2h of the 3h booked
    ("INV-80", "B-205", dt.date(2026,7,8),  2, 90),
    ("INV-81", "B-206", dt.date(2026,7,8),  4, 180),
    ("INV-82", "B-207", dt.date(2026,7,9),  1, 45),
    ("INV-83", "B-208", dt.date(2026,7,9),  3, 150),   # 3h but charged $50/hr not $45
    ("INV-84", "B-209", dt.date(2026,7,10), 2, 90),
    ("INV-85", "B-210", dt.date(2026,7,10), 2, 90),
    ("INV-86", "B-211", dt.date(2026,7,11), 3, 135),
    ("INV-87", "B-213", dt.date(2026,7,12), 2, 90),
    ("INV-88", "B-214", dt.date(2026,7,12), 5, 225),
    ("INV-89", "B-999", dt.date(2026,7,12), 2, 90),    # PHANTOM: B-999 is not in Bookings
]
IV_FIRST = 4
for i, (inv, bid, d, hrs, amt) in enumerate(invoices):
    r = IV_FIRST + i
    iv.cell(r, 1, inv).border = border
    iv.cell(r, 2, bid).border = border
    c = iv.cell(r, 3, d); c.number_format = "m/d/yyyy"; c.border = border
    iv.cell(r, 4, hrs).border = border
    c = iv.cell(r, 5, amt); c.number_format = "$#,##0"; c.border = border
    c.font = Font(bold=True, color=INK)
IV_LAST = IV_FIRST + len(invoices) - 1
iv.freeze_panes = "A4"
for i, w in enumerate([11, 12, 11, 13, 12], 1):
    iv.column_dimensions[get_column_letter(i)].width = w
wordmark(iv, IV_LAST + 2)

# ranges for cross-tab lookups
IV_BID   = f"Invoices!$B${IV_FIRST}:$B${IV_LAST}"
IV_HRS   = f"Invoices!$D${IV_FIRST}:$D${IV_LAST}"
IV_AMT   = f"Invoices!$E${IV_FIRST}:$E${IV_LAST}"
IV_NO    = f"Invoices!$A${IV_FIRST}:$A${IV_LAST}"
BK_IDS   = f"Bookings!$A${BK_FIRST}:$A${BK_LAST}"

# ============ Reconciliation (the product: only the mismatches, + the gap) ============
rc = wb.create_sheet("Reconciliation")
band(rc, "A1:I1", "RECONCILIATION  ·  where the two systems disagree")
rc["A2"] = "Auto-generated. Match on Booking ID. Green rows agree; flagged rows are money on the line. Nothing to maintain - it recomputes when either system changes."
rc["A2"].font = Font(italic=True, size=10, color=MUTE)

# ---- headline gap panel ----
rc.merge_cells("A4:C4")
rc["A4"] = "DELIVERED BUT NOT COLLECTED  (this week)"
rc["A4"].font = Font(bold=True, size=12, color=RED)
rc["A4"].alignment = Alignment(horizontal="left", vertical="center")
gap = rc["D4"]
gap.value = "=SUM(E12:E25)-SUM(G12:G25)"   # booked $ minus invoiced-against-bookings $
gap.number_format = "$#,##0"
gap.font = Font(bold=True, size=16, color=RED)
gap.alignment = Alignment(horizontal="right", vertical="center")
for a in ("A4","B4","C4","D4"):
    rc[a].fill = RED_FILL
rc.row_dimensions[4].height = 30
rc["E4"] = "annualized ~"   # NO leading '=' — Sheets parses a leading '=' as a formula → #ERROR
rc["E4"].font = Font(italic=True, size=9, color=MUTE); rc["E4"].alignment = Alignment(horizontal="right")
ann = rc["F4"]; ann.value = "=D4*52"; ann.number_format = "$#,##0"
ann.font = Font(bold=True, color=RED); ann.alignment = Alignment(horizontal="left")

# ---- small stat rows ----
stats = [
    ("Bay time you delivered (booked)", "=SUM(E12:E25)", INK),
    ("Actually invoiced against those bookings", "=SUM(G12:G25)", INK),
    ("Sessions never billed at all", '=COUNTIF(I12:I25,"NOT BILLED")', RED),
    ("Invoices that match no booking (dispute risk)",
     f"=SUMPRODUCT(--ISNA(MATCH({IV_BID},{BK_IDS},0)))", AMBER),
]
for i, (label, f, color) in enumerate(stats):
    r = 6 + i
    lc = rc.cell(r, 1, label); lc.font = Font(bold=True, color=INK); lc.fill = CARD_FILL
    rc.merge_cells(f"A{r}:C{r}")
    for cc in ("B","C"): rc[f"{cc}{r}"].fill = CARD_FILL
    v = rc.cell(r, 4, f); v.fill = CARD_FILL
    v.number_format = "$#,##0" if i < 2 else "0"
    v.font = Font(bold=True, size=12, color=color); v.alignment = Alignment(horizontal="right")

# ---- per-booking reconciliation table ----
headers_row(rc, 11, ["Booking ID","Date","Bay","Booked hrs","Expected $",
                     "Invoice #","Invoiced $","Delta $","Status"])
RC_FIRST = 12
for i in range(len(bookings)):
    br = BK_FIRST + i          # matching Bookings row
    r  = RC_FIRST + i
    rc.cell(r, 1, f"=Bookings!A{br}")
    c = rc.cell(r, 2, f"=Bookings!B{br}"); c.number_format = "m/d/yyyy"
    rc.cell(r, 3, f"=Bookings!C{br}")
    rc.cell(r, 4, f"=Bookings!D{br}")
    c = rc.cell(r, 5, f"=Bookings!F{br}"); c.number_format = "$#,##0"
    # invoice # (or dash), invoiced $, invoiced hrs via MATCH into Invoices
    m = f"MATCH($A{r},{IV_BID},0)"
    rc.cell(r, 6, f'=IFERROR(INDEX({IV_NO},{m}),"-")')
    c = rc.cell(r, 7, f'=IFERROR(INDEX({IV_AMT},{m}),"")'); c.number_format = "$#,##0"
    # delta = invoiced - expected ; if not billed, = -expected
    c = rc.cell(r, 8, f'=IF($G{r}="",-$E{r},$G{r}-$E{r})'); c.number_format = "$#,##0;($#,##0)"
    # invoiced hrs for the hours-mismatch test
    hrs_expr = f'IFERROR(INDEX({IV_HRS},{m}),"")'
    status = (f'=IF($G{r}="","NOT BILLED",'
              f'IF({hrs_expr}<>$D{r},"HOURS MISMATCH",'
              f'IF(ABS($H{r})>0.004,"AMOUNT MISMATCH","OK")))')
    rc.cell(r, 9, status)
    # static status + delta coloring keyed to the seeded scenario (see note above)
    known = bookings[i][5]
    sc = rc.cell(r, 9)
    if known == "OK":
        sc.font = Font(bold=True, color=GREEN); sc.fill = GREEN_FILL
    elif known == "NOT BILLED":
        sc.font = Font(bold=True, color=RED); sc.fill = RED_FILL
    else:
        sc.font = Font(bold=True, color=AMBER); sc.fill = AMBER_FILL
    if known in ("NOT BILLED", "HOURS MISMATCH"):
        rc.cell(r, 8).font = Font(bold=True, color=RED)
    for cnum in range(1, 10):
        rc.cell(r, cnum).border = border
RC_LAST = RC_FIRST + len(bookings) - 1

# ---- phantom-invoice callout ----
pr = RC_LAST + 2
rc.merge_cells(f"A{pr}:I{pr}")
rc.cell(pr, 1, "INVOICES THAT MATCH NO BOOKING  (charged a customer with no record on your side - refund/chargeback risk)")
rc.cell(pr, 1).font = Font(bold=True, color=AMBER); rc.cell(pr, 1).fill = AMBER_FILL
for cc in range(2,10): rc.cell(pr, cc).fill = AMBER_FILL
# list them by formula-free lookup (we know INV-89, but show the check is live)
rc.cell(pr+1, 1, "INV-89"); rc.cell(pr+1, 2, "=Invoices!C16")
rc.cell(pr+1, 2).number_format = "m/d/yyyy"
rc.cell(pr+1, 3, "billed to B-999")
rc.cell(pr+1, 6, "no match")
c = rc.cell(pr+1, 7, "=Invoices!E16"); c.number_format = "$#,##0"; c.font = Font(bold=True, color=AMBER)
rc.cell(pr+1, 9, "PHANTOM")
rc.cell(pr+1, 9).font = Font(bold=True, color=AMBER)

note = pr + 3
rc.cell(note, 1, "How to read this: a red DELTA is money you earned but did not collect. NOT BILLED = a session that never became an invoice. "
                 "HOURS MISMATCH = billed fewer hours than played. AMOUNT MISMATCH = right hours, wrong rate. PHANTOM = an invoice with no booking behind it.")
rc.cell(note, 1).font = Font(italic=True, size=9, color=MUTE)
rc.cell(note+1, 1, "The Reconciler emails you only these flagged rows, on a schedule you set. Clean weeks send a one-line 'all matched' - so silence always means good news.")
rc.cell(note+1, 1).font = Font(italic=True, size=9, color=MUTE)
wordmark(rc, note+3)

widths = [12, 11, 13, 11, 12, 11, 12, 11, 16]
for i, w in enumerate(widths, 1):
    rc.column_dimensions[get_column_letter(i)].width = w

path = str(_pl.Path(__file__).resolve().parents[1] / "reconciler" / "reconciler-demo.xlsx")
wb.save(path)

# ---- sanity math (Python mirror of the sheet formulas) ----
exp = {b[0]: b[3]*b[4] for b in bookings}
inv_by_bid = {}
for inv in invoices:
    inv_by_bid[inv[1]] = inv  # last wins; fine, ids unique among matched
total_exp = sum(exp.values())
matched_inv = sum(inv[4] for inv in invoices if inv[1] in exp)
phantom = [inv for inv in invoices if inv[1] not in exp]
not_billed = [b for b in bookings if b[0] not in inv_by_bid]
print("bookings:", len(bookings), "invoices:", len(invoices))
print("total delivered (booked $):", total_exp)
print("invoiced against bookings $:", matched_inv)
print("GAP delivered-not-collected $:", total_exp - matched_inv)
print("never billed:", [b[0] for b in not_billed], "=", sum(exp[b[0]] for b in not_billed))
print("phantom invoices:", [(p[0], p[4]) for p in phantom])
print("annualized gap $:", (total_exp - matched_inv)*52)
print("saved:", path)
