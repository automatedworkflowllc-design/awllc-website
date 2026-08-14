# Build "Property Owner Statement Demo - (Automated Workflow)" as xlsx for Drive conversion.
# The product: a property manager logs rent collected + expenses per unit/owner; the
# dashboard auto-generates a per-owner statement (income, expenses, mgmt fee, net payout)
# and flags owners whose net payout DROPPED vs last month (the "explain this before they call" signal).
# Vertical #3 (property management). $650 Dashboard tier. Static colors (imports clean).
import datetime as dt
import pathlib as _pl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

import sys as _sys, os as _os
_sys.path.insert(0, _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "..", "_brand"))
from awllc_brand import INK, GREEN, AMBER, RED, MUTE, CARD, BANDBG, WORDMARK, GREENBG, REDBG, AMBERBG  # single source of truth
HEADER_FILL=PatternFill("solid",fgColor=INK); CARD_FILL=PatternFill("solid",fgColor=CARD)
GREEN_FILL=PatternFill("solid",fgColor=GREENBG); RED_FILL=PatternFill("solid",fgColor=REDBG); AMBER_FILL=PatternFill("solid",fgColor=AMBERBG)
BAND_FILL=PatternFill("solid",fgColor=BANDBG)
thin=Side(style="thin",color="E4DFD1"); border=Border(bottom=thin)

def band(ws,rng,text,size=14):
    ws.merge_cells(rng); top=rng.split(":")[0]
    ws[top]=text; ws[top].font=Font(bold=True,size=size,color="FFFFFF"); ws[top].alignment=Alignment(horizontal="left",vertical="center")
    from openpyxl.utils.cell import range_boundaries
    minc,minr,maxc,maxr=range_boundaries(rng)
    for r in range(minr,maxr+1):
        for c in range(minc,maxc+1): ws.cell(r,c).fill=HEADER_FILL
    ws.row_dimensions[minr].height=30

def wordmark(ws,row):
    ws.cell(row,1,"Built by Automated Workflow  ·  automatedworkflowllc.com  ·  Owner Statement Tracker")
    ws.cell(row,1).font=Font(italic=True,size=9,color=WORDMARK)

def headers_row(ws,r,cols):
    for c,h in enumerate(cols,1):
        cell=ws.cell(row=r,column=c,value=h); cell.font=Font(bold=True,color="FFFFFF"); cell.fill=HEADER_FILL
    ws.row_dimensions[r].height=22

MGMT_FEE_PCT = 0.08  # 8% management fee, standard-ish for the demo

# ============ Rent Roll (office logs rent + expenses per unit) ============
rr=wb.active; rr.title="Rent Roll"
band(rr,"A1:H1","RENT ROLL  (log rent collected + expenses per unit, this month)")
rr["A2"]="One row per unit. Mgmt fee = 8% of rent collected. Net to owner = Rent − Expenses − Mgmt fee. The Owner Statements tab rolls this up per owner automatically."
rr["A2"].font=Font(italic=True,size=10,color=MUTE)
headers_row(rr,3,["Unit","Owner","Address","Rent Due","Rent Collected","Expenses","Last Mo. Net","Notes"])

# (Unit, Owner, Address, RentDue, RentCollected, Expenses, LastMoNet, Notes, flag)
units=[
    ("101","Ardsley Holdings","412 Ardsley Ln",1450,1450, 180,1176,"Routine turnover clean","STABLE"),
    ("102","Ardsley Holdings","414 Ardsley Ln",1500,1500,  90,1289,"","STABLE"),
    ("103","Ardsley Holdings","416 Ardsley Ln",1400,   0, 220,1157,"Vacant — tenant moved out 7/1","DROPPED"),
    ("201","Kestrel Property Trust","88 Kestrel Row",             1250,1250, 650,1088,"AC repair","DROPPED"),
    ("202","Kestrel Property Trust","90 Kestrel Row",              1300,1300, 110,1176,"","STABLE"),
    ("203","Kestrel Property Trust","92 Kestrel Row",              1350,1350, 130,1218,"","STABLE"),
    ("301","Fenmore Rentals","2210 Fenmore Ave",          1600,1600, 950,1435,"Water heater replacement","DROPPED"),
    ("302","Fenmore Rentals","2212 Fenmore Ave",          1550,1550, 140,1367,"","STABLE"),
    ("303","Fenmore Rentals","2214 Fenmore Ave",          1500,1500, 100,1340,"","STABLE"),
    ("401","Ironwood Commercial","4500 Ironwood Blvd",     2100,2100, 260,1700,"","STABLE"),
    ("402","Ironwood Commercial","4502 Ironwood Blvd",     1950,1950, 180,1450,"","STABLE"),
    ("501","Redbrick Partners","1120 Redbrick Way",      1750,1750,3200,1585,"Roof leak repair","DROPPED"),
]

# Build-time gate: this data is served publicly, so it may not name a real
# business from the outreach tracker. See awllc_brand.assert_no_real_contacts.
from awllc_brand import assert_no_real_contacts
assert_no_real_contacts(units, 'the owner statement demo')
RR_FIRST=4
for i,(unit,owner,addr,due,coll,exp,lastnet,notes,flag) in enumerate(units):
    r=RR_FIRST+i
    rr.cell(r,1,unit).border=border
    rr.cell(r,2,owner).border=border
    rr.cell(r,3,addr).border=border
    c=rr.cell(r,4,due); c.number_format="$#,##0"; c.border=border
    c=rr.cell(r,5,coll); c.number_format="$#,##0"; c.border=border
    if coll==0: c.font=Font(bold=True,color=RED)
    c=rr.cell(r,6,exp); c.number_format="$#,##0"; c.border=border
    c=rr.cell(r,7,lastnet); c.number_format="$#,##0"; c.border=border
    rr.cell(r,8,notes).border=border
RR_LAST=RR_FIRST+len(units)-1
rr.freeze_panes="A4"
for i,w in enumerate([8,24,26,10,14,11,13,30],1): rr.column_dimensions[get_column_letter(i)].width=w
wordmark(rr,RR_LAST+2)

R_UNIT=f"'Rent Roll'!$A${RR_FIRST}:$A${RR_LAST}"; R_OWNER=f"'Rent Roll'!$B${RR_FIRST}:$B${RR_LAST}"
R_COLL=f"'Rent Roll'!$E${RR_FIRST}:$E${RR_LAST}"; R_EXP=f"'Rent Roll'!$F${RR_FIRST}:$F${RR_LAST}"
R_LASTNET=f"'Rent Roll'!$G${RR_FIRST}:$G${RR_LAST}"

owners=["Ardsley Holdings","Kestrel Property Trust","Fenmore Rentals","Ironwood Commercial","Redbrick Partners"]

# ============ Owner Statements ============
os_ws=wb.create_sheet("Owner Statements")
band(os_ws,"A1:F1","OWNER STATEMENTS  ·  this month, by owner")
os_ws["A2"]=f"Auto-updates from Rent Roll. Mgmt fee = {int(MGMT_FEE_PCT*100)}% of rent collected. Net payout compares to last month — red means explain it before they call."
os_ws["A2"].font=Font(italic=True,size=10,color=MUTE)

kpis=[
    ("Total rent collected (all owners)", f"=SUM({R_COLL})", INK),
    ("Total expenses", f"=SUM({R_EXP})", INK),
    ("Total mgmt fees earned", f"=SUM({R_COLL})*{MGMT_FEE_PCT}", GREEN),
    ("Total net paid to owners", f"=SUM({R_COLL})-SUM({R_EXP})-SUM({R_COLL})*{MGMT_FEE_PCT}", INK),
]
for i,(label,f,color) in enumerate(kpis):
    r=4+i
    lc=os_ws.cell(r,1,label); lc.font=Font(bold=True,color=INK); lc.fill=CARD_FILL; os_ws.merge_cells(f"A{r}:C{r}")
    for cc in ("B","C"): os_ws[f"{cc}{r}"].fill=CARD_FILL
    v=os_ws.cell(r,4,f); v.number_format="$#,##0"; v.fill=CARD_FILL; v.font=Font(bold=True,size=13,color=color); v.alignment=Alignment(horizontal="right")

band(os_ws,"A9:F9","PER-OWNER STATEMENT",size=12)
headers_row(os_ws,10,["Owner","Rent Collected","Expenses","Mgmt Fee (8%)","Net Payout","vs Last Mo."])
OS_FIRST=11
for i,owner in enumerate(owners):
    r=OS_FIRST+i
    os_ws.cell(r,1,owner).font=Font(bold=True,color=INK)
    c=os_ws.cell(r,2,f'=SUMIF({R_OWNER},A{r},{R_COLL})'); c.number_format="$#,##0"
    c=os_ws.cell(r,3,f'=SUMIF({R_OWNER},A{r},{R_EXP})'); c.number_format="$#,##0"
    c=os_ws.cell(r,4,f'=B{r}*{MGMT_FEE_PCT}'); c.number_format="$#,##0"; c.font=Font(color=GREEN)
    c=os_ws.cell(r,5,f'=B{r}-C{r}-D{r}'); c.number_format="$#,##0;($#,##0)"; c.font=Font(bold=True,color=INK)
    c=os_ws.cell(r,6,f'=E{r}-SUMIF({R_OWNER},A{r},{R_LASTNET})'); c.number_format="$#,##0;($#,##0)"
    for cn in range(1,7): os_ws.cell(r,cn).border=border
OS_LAST=OS_FIRST+len(owners)-1
tr=OS_LAST+1
os_ws.cell(tr,1,"All owners").font=Font(bold=True,color=INK)
for col in (2,3,4,5,6):
    L=get_column_letter(col); c=os_ws.cell(tr,col,f"=SUM({L}{OS_FIRST}:{L}{OS_LAST})"); c.number_format="$#,##0;($#,##0)"; c.font=Font(bold=True,color=INK)

# color the "vs Last Mo." column by seeded direction (static, matches formula sign)
owner_dropped = {"Ardsley Holdings":True,"Kestrel Property Trust":True,"Fenmore Rentals":True,
                 "Ironwood Commercial":False,"Redbrick Partners":True}
for i,owner in enumerate(owners):
    r=OS_FIRST+i; cell=os_ws.cell(r,6)
    if owner_dropped[owner]:
        cell.font=Font(bold=True,color=RED)
        for cn in range(1,7): os_ws.cell(r,cn).fill=RED_FILL
    else:
        cell.font=Font(bold=True,color=GREEN)
        for cn in range(1,7): os_ws.cell(r,cn).fill=GREEN_FILL

# unit-level flag detail (why the drop happened)
band(os_ws,f"A{tr+3}:F{tr+3}","WHY THE DROPS HAPPENED (unit-level detail)",size=12)
headers_row(os_ws,tr+4,["Unit","Owner","Issue","Rent Collected","Expenses","Net Impact"])
drops=[u for u in units if u[8]=="DROPPED"]
dr=tr+5
for i,(unit,owner,addr,due,coll,exp,lastnet,notes,flag) in enumerate(drops):
    r=dr+i
    os_ws.cell(r,1,unit).border=border
    os_ws.cell(r,2,owner).border=border
    os_ws.cell(r,3,notes if notes else "Rent shortfall").border=border
    c=os_ws.cell(r,4,coll); c.number_format="$#,##0"; c.border=border
    if coll==0: c.font=Font(bold=True,color=RED)
    c=os_ws.cell(r,5,exp); c.number_format="$#,##0"; c.border=border; c.font=Font(bold=True,color=AMBER)
    net_this = coll - exp - coll*MGMT_FEE_PCT
    c=os_ws.cell(r,6,round(net_this-lastnet)); c.number_format="$#,##0;($#,##0)"; c.border=border; c.font=Font(bold=True,color=RED)
dr_last=dr+len(drops)-1

note=dr_last+2
os_ws.cell(note,1,"How to read this: 'vs Last Mo.' in red means that owner's payout dropped — usually a vacancy, a repair, or a rent shortfall. The detail table below shows exactly why, unit by unit, so you have the answer ready before the owner calls asking.")
os_ws.cell(note,1).font=Font(italic=True,size=9,color=MUTE)
os_ws.cell(note+1,1,"This becomes the actual statement you send each owner — filter to their row, attach the PDF, done. No manual math, no digging through the rent roll.")
os_ws.cell(note+1,1).font=Font(italic=True,size=9,color=MUTE)
wordmark(os_ws,note+3)

os_ws.column_dimensions["A"].width=26
for col in "BCDEF": os_ws.column_dimensions[col].width=16
os_ws.sheet_view.showGridLines=False

import os as _os
path=str(_pl.Path(__file__).resolve().parents[1] / "owner-statements" / "owner-statement-tracker-demo.xlsx")
_os.makedirs(_os.path.dirname(path),exist_ok=True)
wb.save(path)

# sanity math
tot_coll=sum(u[4] for u in units); tot_exp=sum(u[5] for u in units); tot_fee=tot_coll*MGMT_FEE_PCT
tot_net=tot_coll-tot_exp-tot_fee
print("units:",len(units),"| total collected:",tot_coll,"| total expenses:",tot_exp,"| total mgmt fee:",round(tot_fee),"| total net to owners:",round(tot_net))
by_owner={}
for u in units:
    o=u[1]; by_owner.setdefault(o,{"coll":0,"exp":0,"lastnet":0})
    by_owner[o]["coll"]+=u[4]; by_owner[o]["exp"]+=u[5]; by_owner[o]["lastnet"]+=u[6]
for o,v in by_owner.items():
    net=v["coll"]-v["exp"]-v["coll"]*MGMT_FEE_PCT
    print(f"  {o}: collected={v['coll']} exp={v['exp']} net={round(net)} vs_last={round(net-v['lastnet'])}")
print("saved:",path)
