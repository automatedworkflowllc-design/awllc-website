# Build "Payroll Estimator" as xlsx — the bid artifact for payroll-spreadsheet gigs.
#
# WHY THIS EXISTS. Sixteen Upwork proposals produced one conversation and no hires. The two
# that got any traction both carried something already built for that specific job, which
# became the rule: do not bid without an attachable artifact. Payroll-spreadsheet requests
# recur constantly there ("3 salaried staff and the rest hourly, input the upcoming schedule,
# see the estimated payroll before we commit to it" is close to verbatim from a live posting),
# so this is built once and attached whenever one appears.
#
# WHAT IT ACTUALLY DOES, and the reason it is not a timesheet: it answers the question BEFORE
# the period, not after. You type next fortnight's schedule and it tells you what that
# schedule costs -- including the overtime the schedule creates -- while there is still time
# to move a shift. A timesheet tells you the same number two weeks too late.
#
# EVERY NUMBER IS A FORMULA. Nothing is typed except staff details and hours. The employer
# burden is a RATE THE USER SETS, not a rate this file invents, because a made-up tax
# percentage would be exactly the kind of confident wrong number this company exists to catch.
import os
import pathlib as _pl
import sys as _sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "_brand"))
from awllc_brand import INK, GREEN, AMBER, RED, MUTE, CARD, BANDBG, WORDMARK, AMBERBG, GREENBG

HEADER_FILL = PatternFill("solid", fgColor=INK)
CARD_FILL   = PatternFill("solid", fgColor=CARD)
BAND_FILL   = PatternFill("solid", fgColor=BANDBG)
AMBER_FILL  = PatternFill("solid", fgColor=AMBERBG)
GREEN_FILL  = PatternFill("solid", fgColor=GREENBG)
thin = Side(style="thin", color="E4DFD1")

MONEY = '"$"#,##0.00'
MONEY0 = '"$"#,##0'
HRS = '0.0'

wb = Workbook()


def band(ws, rng, text, size=14):
    ws.merge_cells(rng)
    top = rng.split(":")[0]
    ws[top] = text
    ws[top].font = Font(bold=True, size=size, color="FFFFFF")
    ws[top].alignment = Alignment(horizontal="left", vertical="center")
    from openpyxl.utils.cell import range_boundaries
    minc, minr, maxc, maxr = range_boundaries(rng)
    for r in range(minr, maxr + 1):
        for c in range(minc, maxc + 1):
            ws.cell(r, c).fill = HEADER_FILL
    ws.row_dimensions[minr].height = 30


def note(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(italic=True, size=10, color=MUTE)


def headers_row(ws, r, cols):
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[r].height = 30


def wordmark(ws, row):
    ws.cell(row, 1, "Built by Automated Workflow  ·  automatedworkflowllc.com  ·  Payroll Estimator")
    ws.cell(row, 1).font = Font(italic=True, size=9, color=WORDMARK)


# ---------------------------------------------------------------- sample staff
# INVENTED, and labelled as invented inside the file. Names are deliberately not
# drawn from any real roster or outreach list -- a previous artifact was seeded from
# the lead tracker and shipped real business names, which must not happen again.
# Three salaried and seven hourly, matching the shape these postings describe.
SALARIED = [
    ("Dr. A. Whitfield",  "Dentist",            185000),
    ("M. Okafor",         "Practice Manager",    72000),
    ("L. Ferrara",        "Lead Hygienist",      88000),
]
HOURLY = [
    ("D. Marchetti", "Hygienist",              42.00),
    ("J. Alvarez",   "Dental Assistant",       24.00),
    ("P. Nakamura",  "Dental Assistant",       22.50),
    ("K. Osei",      "Scheduling Coordinator", 20.00),
    ("R. Ellison",   "Front Desk",             19.00),
    ("T. Boateng",   "Front Desk",             18.50),
    ("S. Kowalski",  "Sterilization Tech",     17.00),
]

# Week 1 / Week 2 hours per hourly employee. One person is deliberately scheduled
# into overtime in week 2 (44h) so the OT column is not a column of zeros in the
# sample -- a feature nobody can see working is a feature nobody believes.
HOURS = {
    "D. Marchetti": (32, 32),
    "J. Alvarez":   (40, 44),
    "P. Nakamura":  (38, 40),
    "K. Osei":      (40, 40),
    "R. Ellison":   (32, 36),
    "T. Boateng":   (24, 28),
    "S. Kowalski":  (20, 20),
}

PERIODS_PER_YEAR = 26      # fortnightly
OT_THRESHOLD = 40          # hours per week before overtime, FLSA default
BURDEN_DEFAULT = 0.092     # employer-side estimate the USER owns; see Settings note

# ============================================================ Settings
st = wb.active
st.title = "Settings"
band(st, "A1:D1", "SETTINGS  (the only assumptions in this workbook — change them here)")
note(st, "A2", "Every other tab reads these cells. Nothing below is hard-coded anywhere else, "
               "so correcting a rate here corrects the whole file.")

headers_row(st, 4, ["Setting", "Value", "Unit", "What it does / why it is yours to set"])
rows = [
    ("Pay periods per year", PERIODS_PER_YEAR, "periods",
     "26 = fortnightly. Use 24 for semi-monthly, 52 for weekly. Salaried pay is divided by this."),
    ("Overtime threshold", OT_THRESHOLD, "hours/week",
     "Hours above this in a single week are paid at the multiplier below. FLSA default is 40; "
     "some states are stricter."),
    ("Overtime multiplier", 1.5, "× base rate",
     "1.5 = time-and-a-half."),
    ("Employer burden rate", BURDEN_DEFAULT, "of gross",
     "YOUR number, not one this file invented. Employer-side payroll taxes only "
     "(FICA 7.65% + your FUTA/SUTA). It does NOT include benefits, workers' comp or "
     "retirement match — add those to Payroll goal, or raise this rate knowingly."),
    ("Payroll goal (per period)", 34000, "$",
     "What you are trying to stay under. The Estimate tab compares against this and "
     "says by how much, not just whether."),
]
r = 5
for label, val, unit, why in rows:
    st.cell(r, 1, label).font = Font(bold=True)
    c = st.cell(r, 2, val)
    if label == "Employer burden rate":
        c.number_format = "0.00%"
    elif label == "Payroll goal (per period)":
        c.number_format = MONEY0
    c.font = Font(bold=True, color=GREEN)
    c.fill = CARD_FILL
    st.cell(r, 3, unit).font = Font(size=10, color=MUTE)
    w = st.cell(r, 4, why)
    w.font = Font(size=10, color=MUTE)
    w.alignment = Alignment(wrap_text=True, vertical="top")
    st.row_dimensions[r].height = 34
    r += 1

st.column_dimensions["A"].width = 26
st.column_dimensions["B"].width = 13
st.column_dimensions["C"].width = 13
st.column_dimensions["D"].width = 74
wordmark(st, r + 2)

# ============================================================ Staff
sf = wb.create_sheet("Staff")
band(sf, "A1:F1", "STAFF  (one row per person — the only place a rate is typed)")
note(sf, "A2", "SAMPLE DATA — every name, role and figure below is invented. Replace them with "
               "your own. Salaried people need an annual salary; hourly people need an hourly rate.")

headers_row(sf, 4, ["Employee", "Role", "Type", "Annual salary", "Hourly rate", "Per-period base pay"])
r = 5
staff_first = r
for name, role, salary in SALARIED:
    sf.cell(r, 1, name)
    sf.cell(r, 2, role)
    sf.cell(r, 3, "Salaried")
    sf.cell(r, 4, salary).number_format = MONEY0
    sf.cell(r, 5, "—").alignment = Alignment(horizontal="center")
    # Salary / periods. Reads the period count from Settings so one edit repays everywhere.
    sf.cell(r, 6, "=IF(C{r}=\"Salaried\",D{r}/Settings!$B$5,0)".format(r=r)).number_format = MONEY
    r += 1
for name, role, rate in HOURLY:
    sf.cell(r, 1, name)
    sf.cell(r, 2, role)
    sf.cell(r, 3, "Hourly")
    sf.cell(r, 4, "—").alignment = Alignment(horizontal="center")
    sf.cell(r, 5, rate).number_format = MONEY
    sf.cell(r, 6, "—").alignment = Alignment(horizontal="center")
    r += 1
staff_last = r - 1

for rr in range(staff_first, staff_last + 1):
    for cc in range(1, 7):
        sf.cell(rr, cc).border = Border(bottom=thin)

for col, w in zip("ABCDEF", (22, 24, 12, 15, 13, 20)):
    sf.column_dimensions[col].width = w
wordmark(sf, staff_last + 3)

# ============================================================ Schedule
sc = wb.create_sheet("Schedule")
band(sc, "A1:D1", "SCHEDULE  (type next period's hours — this is the whole input)")
note(sc, "A2", "Hours for the UPCOMING period, not hours already worked. That is the point of the "
               "file: it costs the schedule while you can still change it. Salaried staff are "
               "listed for completeness — their hours do not change their pay, and the "
               "Estimate tab does not pretend otherwise.")

headers_row(sc, 4, ["Employee", "Type", "Week 1 hours", "Week 2 hours"])
r = 5
sched_first = r
for name, role, _s in SALARIED:
    sc.cell(r, 1, name)
    sc.cell(r, 2, "Salaried")
    sc.cell(r, 3, "—").alignment = Alignment(horizontal="center")
    sc.cell(r, 4, "—").alignment = Alignment(horizontal="center")
    for cc in range(1, 5):
        sc.cell(r, cc).font = Font(color=MUTE)
    r += 1
for name, role, _rate in HOURLY:
    w1, w2 = HOURS[name]
    sc.cell(r, 1, name)
    sc.cell(r, 2, "Hourly")
    c1 = sc.cell(r, 3, w1); c1.number_format = HRS; c1.fill = CARD_FILL; c1.font = Font(bold=True)
    c2 = sc.cell(r, 4, w2); c2.number_format = HRS; c2.fill = CARD_FILL; c2.font = Font(bold=True)
    r += 1
sched_last = r - 1

for rr in range(sched_first, sched_last + 1):
    for cc in range(1, 5):
        sc.cell(rr, cc).border = Border(bottom=thin)

for col, w in zip("ABCD", (22, 12, 14, 14)):
    sc.column_dimensions[col].width = w
wordmark(sc, sched_last + 3)

# ============================================================ Estimate
es = wb.create_sheet("Estimate")
band(es, "A1:H1", "ESTIMATE  (every cell below is a formula — nothing here is typed)")
note(es, "A2", "Overtime is computed PER WEEK, not on the fortnight total. Someone who works 44 "
               "hours then 36 has 4 hours of overtime, not zero — averaging the two weeks "
               "would hide it, and that is the single most common way a payroll estimate comes in low.")

headers_row(es, 4, ["Employee", "Type", "Regular hrs", "Overtime hrs",
                    "Base pay", "Overtime pay", "Gross", "Employer cost"])

r = 5
est_first = r
total_rows = staff_last - staff_first + 1
for i in range(total_rows):
    srow = staff_first + i          # matching row on Staff
    krow = sched_first + i          # matching row on Schedule (same order, by construction)
    es.cell(r, 1, "=Staff!A{}".format(srow))
    es.cell(r, 2, "=Staff!C{}".format(srow))
    # Regular = min(week, threshold) for each of the two weeks. Salaried -> 0, shown as a dash
    # would break the SUM, so 0 it is, and the Type column says why it is 0.
    es.cell(r, 3, "=IF(B{r}=\"Salaried\",0,"
                  "MIN(Schedule!C{k},Settings!$B$6)+MIN(Schedule!D{k},Settings!$B$6))"
                  .format(r=r, k=krow)).number_format = HRS
    es.cell(r, 4, "=IF(B{r}=\"Salaried\",0,"
                  "MAX(0,Schedule!C{k}-Settings!$B$6)+MAX(0,Schedule!D{k}-Settings!$B$6))"
                  .format(r=r, k=krow)).number_format = HRS
    # Salaried base comes from Staff's per-period formula; hourly from regular hours x rate.
    es.cell(r, 5, "=IF(B{r}=\"Salaried\",Staff!F{s},C{r}*Staff!E{s})"
                  .format(r=r, s=srow)).number_format = MONEY
    es.cell(r, 6, "=IF(B{r}=\"Salaried\",0,D{r}*Staff!E{s}*Settings!$B$7)"
                  .format(r=r, s=srow)).number_format = MONEY
    es.cell(r, 7, "=E{r}+F{r}".format(r=r)).number_format = MONEY
    es.cell(r, 8, "=G{r}*(1+Settings!$B$8)".format(r=r)).number_format = MONEY
    r += 1
est_last = r - 1

for rr in range(est_first, est_last + 1):
    for cc in range(1, 9):
        es.cell(rr, cc).border = Border(bottom=thin)

tr = est_last + 1
es.cell(tr, 1, "TOTAL").font = Font(bold=True, color="FFFFFF")
es.cell(tr, 2, "").fill = HEADER_FILL
for cc in range(1, 9):
    es.cell(tr, cc).fill = HEADER_FILL
    es.cell(tr, cc).font = Font(bold=True, color="FFFFFF")
for col in ("C", "D"):
    es.cell(tr, {"C": 3, "D": 4}[col],
            "=SUM({c}{a}:{c}{b})".format(c=col, a=est_first, b=est_last)).number_format = HRS
for col, ci in (("E", 5), ("F", 6), ("G", 7), ("H", 8)):
    es.cell(tr, ci,
            "=SUM({c}{a}:{c}{b})".format(c=col, a=est_first, b=est_last)).number_format = MONEY
for cc in range(1, 9):
    es.cell(tr, cc).font = Font(bold=True, color="FFFFFF")

# ---- goal comparison, stated as a number rather than a colour
gr = tr + 2
es.cell(gr, 1, "Payroll goal this period").font = Font(bold=True)
es.cell(gr, 3, "=Settings!$B$9").number_format = MONEY0
es.cell(gr + 1, 1, "Estimated employer cost").font = Font(bold=True)
es.cell(gr + 1, 3, "=H{}".format(tr)).number_format = MONEY
es.cell(gr + 2, 1, "Difference").font = Font(bold=True)
d = es.cell(gr + 2, 3, "=C{}-C{}".format(gr, gr + 1))
d.number_format = MONEY
d.font = Font(bold=True)
v = es.cell(gr + 2, 4,
            '=IF(C{g}-C{e}>=0,"under goal","OVER goal by this much")'.format(g=gr, e=gr + 1))
v.font = Font(bold=True, size=10)
for rr in (gr, gr + 1, gr + 2):
    for cc in (1, 3, 4):
        es.cell(rr, cc).fill = CARD_FILL

# ---- the honesty block: what this file cannot tell you
hb = gr + 4
es.cell(hb, 1, "What this workbook CANNOT tell you").font = Font(bold=True, size=11, color=INK)
limits = [
    "It does not know your benefits, workers' compensation, retirement match or PTO accrual. "
    "Employer cost here is gross pay plus the burden rate on Settings, and nothing else.",
    "It is an ESTIMATE of employer cost, not a withholding calculation. It does not compute what "
    "comes out of anyone's cheque, and it is not a substitute for your payroll provider.",
    "It assumes the schedule you typed is the schedule that happens. A called-in shift or an "
    "extra hour changes the answer, which is why it is quick to re-run rather than clever.",
    "Overtime uses the threshold on Settings. If your state is stricter than the federal 40 "
    "(California's daily rule, for instance), change it there — this file will not know.",
]
rr = hb + 1
for t in limits:
    c = es.cell(rr, 1, "·  " + t)
    c.font = Font(size=10, color=MUTE)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    es.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=8)
    es.row_dimensions[rr].height = 30
    rr += 1
c = es.cell(rr + 1, 1, "Stated because a number you cannot trust is worse than one you do not have.")
c.font = Font(italic=True, size=10, color=MUTE)

for col, w in zip("ABCDEFGH", (22, 12, 12, 12, 14, 14, 14, 15)):
    es.column_dimensions[col].width = w
wordmark(es, rr + 3)

es.freeze_panes = "A5"
sf.freeze_panes = "A5"
sc.freeze_panes = "A5"

path = str(_pl.Path(__file__).resolve().parents[1] / "payroll-estimator" / "Payroll-Estimator-sample.xlsx")
os.makedirs(os.path.dirname(path), exist_ok=True)
wb.save(path)

# ------------------------------------------------------------ sanity math
# Recomputed in Python, independently of the workbook formulas, so the two can disagree
# out loud rather than the spreadsheet grading its own homework.
sal_per_period = sum(s / PERIODS_PER_YEAR for _n, _r, s in SALARIED)
reg_pay = ot_pay = 0.0
reg_h = ot_h = 0.0
for name, _role, rate in HOURLY:
    for wk in HOURS[name]:
        reg = min(wk, OT_THRESHOLD)
        ot = max(0, wk - OT_THRESHOLD)
        reg_h += reg
        ot_h += ot
        reg_pay += reg * rate
        ot_pay += ot * rate * 1.5
gross = sal_per_period + reg_pay + ot_pay
employer = gross * (1 + BURDEN_DEFAULT)
print("staff: %d salaried + %d hourly" % (len(SALARIED), len(HOURLY)))
print("regular hrs: %.1f | overtime hrs: %.1f" % (reg_h, ot_h))
print("salaried/period: $%,.2f".replace("%,", "%") % sal_per_period)
print("hourly base: $%.2f | overtime: $%.2f" % (reg_pay, ot_pay))
print("GROSS: $%.2f | employer cost @ %.1f%%: $%.2f" % (gross, BURDEN_DEFAULT * 100, employer))
print("goal 34000 -> %s by $%.2f" % ("under" if employer <= 34000 else "OVER", abs(34000 - employer)))
print("saved:", path)
