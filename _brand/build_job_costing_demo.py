# Build "Job Costing Tracker Demo - (Automated Workflow)" as xlsx for Drive conversion.
# The product: a contractor logs quoted price + actual labor/materials per job; the
# dashboard shows margin per job and flags which jobs are BLEEDING money (actual > quoted).
# Vertical #2 (contractor job-costing) — maps to our real outreach list (lawn/fence/paint/clean/pest).
# $650 Dashboard-tier. Static colors (no conditional formatting → imports clean into Sheets).
import datetime as dt
import pathlib as _pl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

import sys as _sys, os as _os
_sys.path.insert(0, _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "..", "_brand"))
from awllc_brand import INK, GREEN, AMBER, RED, MUTE, CARD, BANDBG, WORDMARK, GREENBG, REDBG, AMBERBG  # single source of truth
HEADER_FILL=PatternFill("solid",fgColor=INK); CARD_FILL=PatternFill("solid",fgColor=CARD)
GREEN_FILL=PatternFill("solid",fgColor=GREENBG); RED_FILL=PatternFill("solid",fgColor=REDBG); AMBER_FILL=PatternFill("solid",fgColor=AMBERBG)
BAND_FILL=PatternFill("solid",fgColor=BANDBG)
thin=Side(style="thin",color="E4DFD1"); border=Border(bottom=thin)

def band(ws,rng,text,size=14):
    ws.merge_cells(rng); top=rng.split(":")[0]
    ws[top]=text; ws[top].font=Font(bold=True,size=size,color="FFFFFF"); ws[top].alignment=Alignment(horizontal="left",vertical="center")
    from openpyxl.utils.cell import range_boundaries
    minc,minr,maxc,maxr=range_boundaries(rng)
    for r in range(minr,maxr+1):
        for c in range(minc,maxc+1): ws.cell(r,c).fill=HEADER_FILL
    ws.row_dimensions[minr].height=30

def wordmark(ws,row):
    ws.cell(row,1,"Built by Automated Workflow  ·  automatedworkflowllc.com  ·  Job Costing Tracker")
    ws.cell(row,1).font=Font(italic=True,size=9,color=WORDMARK)

def headers_row(ws,r,cols):
    for c,h in enumerate(cols,1):
        cell=ws.cell(row=r,column=c,value=h); cell.font=Font(bold=True,color="FFFFFF"); cell.fill=HEADER_FILL
    ws.row_dimensions[r].height=22

# ============ Jobs (crew/office types the numbers) ============
jb=wb.active; jb.title="Jobs"
band(jb,"A1:G1","JOBS  (log quoted price + actual labor & materials per job)")
jb["A2"]="One row per job. Actual cost = Labor + Materials. The Job Dashboard tab does the margin math and flags the losers automatically."
jb["A2"].font=Font(italic=True,size=10,color=MUTE)
headers_row(jb,3,["Job ID","Client","Job Type","Quoted $","Labor $","Materials $","Status"])

# (Job ID, Client, Type, Quoted, Labor, Materials, Status, known_flag)
jobs=[
    ("J-01","Cedar Row Landscaping","Lawn install",       3200,1400, 900,"Complete","HEALTHY"),
    ("J-02","Northgate Fence Co.",          "Fence build",        5400,2600,1800,"Complete","HEALTHY"),
    ("J-03","Bellamy Exterior Painting",            "Exterior paint",     4800,2900,2400,"Complete","OVER BUDGET"),
    ("J-04","Two Creeks Lawn Care",          "Sod + irrigation",   2900,1200, 800,"Complete","HEALTHY"),
    ("J-05","Marlow Landscape Design",    "Landscape design",   3600,1800,1100,"Complete","HEALTHY"),
    ("J-06","Brightline Cleaning",        "Deep-clean contract",1800,1500, 200,"Complete","THIN MARGIN"),
    ("J-07","Sawgrass Pest Control",               "Quarterly setup",    1200, 500, 250,"Complete","HEALTHY"),
    ("J-08","Ridgeway Hardscapes",       "Retaining wall",     6800,3800,3600,"Complete","OVER BUDGET"),
    ("J-09","Fenwick Hardware",           "Fixture install",    2400,1100, 700,"Complete","HEALTHY"),
    ("J-10","Quarry Lane Storage",        "Lot resurfacing",    4200,1900,1500,"In Progress","HEALTHY"),
    ("J-11","Trailhead Cycles",     "Deck build",         3100,1600,1200,"In Progress","THIN MARGIN"),
    ("J-12","Keystone Cafe",          "Patio install",      5200,2400,1900,"In Progress","HEALTHY"),
]

# Build-time gate: this data is served publicly, so it may not name a real
# business from the outreach tracker. See awllc_brand.assert_no_real_contacts.
from awllc_brand import assert_no_real_contacts
assert_no_real_contacts(jobs, 'the job costing demo')
JB_FIRST=4
for i,(jid,client,typ,quoted,labor,mat,status,flag) in enumerate(jobs):
    r=JB_FIRST+i
    jb.cell(r,1,jid).border=border
    jb.cell(r,2,client).border=border
    jb.cell(r,3,typ).border=border
    c=jb.cell(r,4,quoted); c.number_format="$#,##0"; c.border=border; c.font=Font(bold=True,color=INK)
    c=jb.cell(r,5,labor); c.number_format="$#,##0"; c.border=border
    c=jb.cell(r,6,mat); c.number_format="$#,##0"; c.border=border
    sc=jb.cell(r,7,status); sc.border=border
    sc.font=Font(color=GREEN) if status=="Complete" else Font(color=AMBER)
JB_LAST=JB_FIRST+len(jobs)-1
jb.freeze_panes="A4"
for i,w in enumerate([9,28,20,11,11,13,12],1): jb.column_dimensions[get_column_letter(i)].width=w
wordmark(jb,JB_LAST+2)

J_ID=f"Jobs!$A${JB_FIRST}:$A${JB_LAST}"; J_QUOTED=f"Jobs!$D${JB_FIRST}:$D${JB_LAST}"
J_LABOR=f"Jobs!$E${JB_FIRST}:$E${JB_LAST}"; J_MAT=f"Jobs!$F${JB_FIRST}:$F${JB_LAST}"

# ============ Job Dashboard ============
db=wb.create_sheet("Job Dashboard")
band(db,"A1:F1","JOB DASHBOARD  ·  which jobs make money, which bleed it")
db["A2"]="Auto-updates from Jobs. Margin = Quoted − (Labor + Materials). Red = you lost money on it; amber = margin under 15%."
db["A2"].font=Font(italic=True,size=10,color=MUTE)

# headline: money left on jobs that went over
db.merge_cells("A4:C4")
db["A4"]="MONEY LOST ON OVER-BUDGET JOBS"
db["A4"].font=Font(bold=True,size=12,color=RED); db["A4"].alignment=Alignment(horizontal="left",vertical="center")
lost=db["D4"]; lost.value="=SUMIF(H12:H23,\"OVER BUDGET\",F12:F23)"
lost.number_format="$#,##0;($#,##0)"; lost.font=Font(bold=True,size=16,color=RED); lost.alignment=Alignment(horizontal="right",vertical="center")
for a in ("A4","B4","C4","D4"): db[a].fill=RED_FILL
db.row_dimensions[4].height=30

stats=[
    ("Total quoted (booked work)", "=SUM(D12:D23)", INK),
    ("Total actual cost (labor + materials)", "=SUM(E12:E23)", INK),
    ("Total margin kept", "=SUM(F12:F23)", GREEN),
    ("Blended margin %", "=SUM(F12:F23)/SUM(D12:D23)", INK),
    ("Jobs over budget / thin (<15%)", "=COUNTIF(H12:H23,\"OVER BUDGET\")+COUNTIF(H12:H23,\"THIN MARGIN\")", AMBER),
]
for i,(label,f,color) in enumerate(stats):
    r=6+i
    lc=db.cell(r,1,label); lc.font=Font(bold=True,color=INK); lc.fill=CARD_FILL; db.merge_cells(f"A{r}:C{r}")
    for cc in ("B","C"): db[f"{cc}{r}"].fill=CARD_FILL
    v=db.cell(r,4,f); v.fill=CARD_FILL; v.font=Font(bold=True,size=12,color=color); v.alignment=Alignment(horizontal="right")
    v.number_format="0.0%" if i==3 else ("0" if i==4 else "$#,##0")

# per-job table
headers_row(db,11,["Job ID","Client","Job Type","Quoted $","Actual cost $","Margin $","Margin %","Flag"])
DB_FIRST=12
for i in range(len(jobs)):
    jr=JB_FIRST+i; r=DB_FIRST+i
    db.cell(r,1,f"=Jobs!A{jr}")
    db.cell(r,2,f"=Jobs!B{jr}")
    db.cell(r,3,f"=Jobs!C{jr}")
    c=db.cell(r,4,f"=Jobs!D{jr}"); c.number_format="$#,##0"
    c=db.cell(r,5,f"=Jobs!E{jr}+Jobs!F{jr}"); c.number_format="$#,##0"   # actual = labor+materials
    c=db.cell(r,6,f"=D{r}-E{r}"); c.number_format="$#,##0;($#,##0)"      # margin
    c=db.cell(r,7,f"=IF(D{r}=0,0,F{r}/D{r})"); c.number_format="0.0%"    # margin %
    flag=jobs[i][7]
    fc=db.cell(r,8,f'=IF(F{r}<0,"OVER BUDGET",IF(F{r}/D{r}<0.15,"THIN MARGIN","HEALTHY"))')
    # static color keyed to seeded flag
    if flag=="OVER BUDGET":
        fc.font=Font(bold=True,color=RED); [setattr(db.cell(r,cn),'fill',RED_FILL) for cn in range(1,9)]
        db.cell(r,6).font=Font(bold=True,color=RED)
    elif flag=="THIN MARGIN":
        fc.font=Font(bold=True,color=AMBER); [setattr(db.cell(r,cn),'fill',AMBER_FILL) for cn in range(1,9)]
    else:
        fc.font=Font(bold=True,color=GREEN); [setattr(db.cell(r,cn),'fill',GREEN_FILL) for cn in range(1,9)]
    for cn in range(1,9): db.cell(r,cn).border=border
DB_LAST=DB_FIRST+len(jobs)-1
# totals row
tr=DB_LAST+1
db.cell(tr,3,"Portfolio total").font=Font(bold=True,color=INK)
for col,fmt in ((4,"$#,##0"),(5,"$#,##0"),(6,"$#,##0;($#,##0)")):
    L=get_column_letter(col); c=db.cell(tr,col,f"=SUM({L}{DB_FIRST}:{L}{DB_LAST})"); c.number_format=fmt; c.font=Font(bold=True,color=INK)
c=db.cell(tr,7,f"=F{tr}/D{tr}"); c.number_format="0.0%"; c.font=Font(bold=True,color=INK)

note=tr+2
db.cell(note,1,"How to read this: Margin $ is what you kept after labor + materials. A red row means the job cost more than you quoted — you paid to do it. Amber = under 15% margin, barely worth the truck roll.")
db.cell(note,1).font=Font(italic=True,size=9,color=MUTE)
db.cell(note+1,1,"Sort by Margin % low-to-high before you quote the next one — your losers are usually the same job type every time. That's where the money is.")
db.cell(note+1,1).font=Font(italic=True,size=9,color=MUTE)
wordmark(db,note+3)
db.column_dimensions["A"].width=9
for col in "BC": db.column_dimensions[col].width=24 if col=="B" else 20
for col in "DEFGH": db.column_dimensions[col].width=13
db.sheet_view.showGridLines=False

import os
path=str(_pl.Path(__file__).resolve().parents[1] / "job-costing" / "job-costing-tracker-demo.xlsx")
os.makedirs(os.path.dirname(path),exist_ok=True)
wb.save(path)

# sanity math
tq=sum(j[3] for j in jobs); ta=sum(j[4]+j[5] for j in jobs); tm=tq-ta
over=[j for j in jobs if (j[3]-(j[4]+j[5]))<0]
print("jobs:",len(jobs),"| total quoted:",tq,"| total actual:",ta,"| margin:",tm,"| blended %:",round(tm/tq*100,1))
print("over-budget:",[(j[0],j[3]-(j[4]+j[5])) for j in over],"= lost",sum(j[3]-(j[4]+j[5]) for j in over))
print("saved:",path)
