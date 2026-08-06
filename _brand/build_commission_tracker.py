#!/usr/bin/env python3
"""Build the free staffing commission tracker workbook.

WHY THIS EXISTS: /free-staffing-commission-tracker/ has been live promising
"a Google Sheet you can start using today... Free template, yours to copy"
and "not a signup wall" -- while containing zero copy links, because the sheet
was specced in July and never built. Even the email path had nothing to send.
A page that promises an artifact which does not exist is the worst defect on
the site; this makes the promise true.

Built from Documents/AWLLC-free-commission-tracker-spec.md (5 tabs, formulas
F1-F8) as a real .xlsx rather than a Drive sheet, because a downloadable file
needs nobody's Google account and no email -- which is what lets the page's
"no signup wall" line stay honest. Opens in Excel, Numbers or Google Sheets.

Excel note: the spec's formulas use Sheets-style open-ended ranges
(Placements!$G$2:$G). Excel rejects those, so every range here is bounded to
LAST_ROW. Changing that bound means changing it in one place.
"""
from __future__ import annotations

import datetime as _dt
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import sys as _sys
_sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from awllc_brand import INK, GREEN, AMBER, RED, MUTE, CARD, BANDBG, WORDMARK  # brand tokens

ROOT = pathlib.Path(__file__).resolve().parents[1]
OUT = ROOT / 'free-staffing-commission-tracker' / 'staffing-commission-tracker.xlsx'

LAST = 501            # bounded ranges: Excel will not take Sheets' open-ended form
HDR_FILL = PatternFill('solid', fgColor=INK)
HDR_FONT = Font(color='FFFFFF', bold=True, size=10)
CALC_FILL = PatternFill('solid', fgColor=CARD)      # computed columns look different
NOTE_FONT = Font(color=MUTE, size=9, italic=True)
THIN = Side(style='thin', color='D9D9D9')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '"$"#,##0'
PCT = '0%'


def header(ws, row, labels, widths=None):
    for i, text in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=text)
        c.fill, c.font, c.border = HDR_FILL, HDR_FONT, BOX
        c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 30
    for i, w in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build() -> pathlib.Path:
    wb = Workbook()

    # ---------- Settings (built first: everything else points at it) --------
    st = wb.active
    st.title = 'Settings'
    st['A1'] = 'Settings — change these and every tab follows'
    st['A1'].font = Font(bold=True, size=13, color=INK)
    st['A3'] = 'Commission rate on gross'
    st['B3'] = 1.0
    st['B3'].number_format = PCT
    st['C3'] = 'House share of the fee/margin before splits. 100% = recruiters split the whole thing.'
    st['C3'].font = NOTE_FONT
    st['A5'] = 'Default owner split'
    st['B5'] = 0.7
    st['B5'].number_format = PCT
    st['C5'] = 'Used when you add a row; override per deal in Placements column P.'
    st['C5'].font = NOTE_FONT

    header(st, 7, ['Placement Types', 'Guarantee Options (days)', 'Statuses'], [22, 26, 22])
    for i, v in enumerate(['Perm', 'Contract'], start=8):
        st.cell(row=i, column=1, value=v).border = BOX
    for i, v in enumerate([30, 60, 90], start=8):
        st.cell(row=i, column=2, value=v).border = BOX
    for i, v in enumerate(['Pending', 'Started', 'Paid', 'Fell off'], start=8):
        st.cell(row=i, column=3, value=v).border = BOX
    st.column_dimensions['C'].width = 70

    # ---------- Recruiters --------------------------------------------------
    rc = wb.create_sheet('Recruiters')
    header(rc, 1,
           ['Recruiter', 'Role', 'Draw / month $', 'Default Owner Split %',
            'Earned MTD $  (auto)', 'Draw Balance $  (auto)'],
           [22, 16, 17, 21, 20, 20])
    seed = [('Alex Rivera', 'Recruiter', 3000, 0.7),
            ('Sam Chen', 'Recruiter', 3000, 0.7),
            ('Jordan Blake', 'Sourcer', 2000, 0.3)]
    # row 12 carries the no-sourcer option so column H's dropdown can offer it
    for r, (name, role, draw, split) in enumerate(seed, start=2):
        rc.cell(row=r, column=1, value=name).border = BOX
        rc.cell(row=r, column=2, value=role).border = BOX
        c = rc.cell(row=r, column=3, value=draw); c.number_format = MONEY; c.border = BOX
        c = rc.cell(row=r, column=4, value=split); c.number_format = PCT; c.border = BOX
    for r in range(2, 12):
        # F7 -- this month's net as owner, plus their sourcer share
        # DELIBERATE DEVIATION FROM THE SPEC. Spec F7 credits the owner with
        # Net (the whole gross) and the sourcer with their share on top, which
        # overstates the owner by their split -- on the sample contract deal
        # that is $24,336 credited where $17,035 was earned. A free template
        # whose headline per-person number is wrong by 43% is worse than none,
        # so the owner is credited with Owner Commission (Q) instead, and both
        # legs exclude deals marked fallen off.
        f = (f'=SUMIFS(Placements!$Q$2:$Q${LAST},Placements!$G$2:$G${LAST},$A{r},'
             f'Placements!$V$2:$V${LAST},"<>Yes",'
             f'Placements!$B$2:$B${LAST},">="&EOMONTH(TODAY(),-1)+1,'
             f'Placements!$B$2:$B${LAST},"<="&EOMONTH(TODAY(),0))'
             f'+SUMIFS(Placements!$R$2:$R${LAST},Placements!$H$2:$H${LAST},$A{r},'
             f'Placements!$V$2:$V${LAST},"<>Yes",'
             f'Placements!$B$2:$B${LAST},">="&EOMONTH(TODAY(),-1)+1,'
             f'Placements!$B$2:$B${LAST},"<="&EOMONTH(TODAY(),0))')
        c = rc.cell(row=r, column=5, value=f)
        c.number_format, c.fill, c.border = MONEY, CALC_FILL, BOX
        c = rc.cell(row=r, column=6, value=f'=E{r}-C{r}')          # F8
        c.number_format, c.fill, c.border = MONEY, CALC_FILL, BOX
    rc.cell(row=12, column=1, value='— none —').border = BOX
    rc['H2'] = 'Draw Balance negative = they have not yet earned the draw back.'
    rc['H2'].font = NOTE_FONT

    # ---------- Placements --------------------------------------------------
    pl = wb.create_sheet('Placements')
    cols = ['Placement ID', 'Date Placed', 'Candidate', 'Client / Account', 'Role Title',
            'Placement Type', 'Recruiter (Owner)', 'Sourcer (Split)',
            'Salary (Perm)', 'Fee % (Perm)', 'Bill Rate /hr', 'Pay Rate /hr',
            'Hours / Week', 'Contract Weeks',
            'Gross Commission $  (auto)', 'Owner Split %', 'Owner Commission $  (auto)',
            'Sourcer Commission $  (auto)', 'Guarantee Days', 'Status', 'Paid Out?',
            'Fell Off?', 'Guarantee Ends  (auto)', 'Clawback Risk?  (auto)',
            'Net Commission $  (auto)']
    header(pl, 1, cols, [13, 12, 18, 20, 18, 15, 18, 18, 13, 11, 12, 12, 12, 13,
                         18, 13, 18, 19, 14, 12, 11, 11, 15, 17, 18])
    pl.freeze_panes = 'A2'

    demo = [
        ('P-001', '2026-06-02', 'Dana Whitfield', 'Northgate Logistics',
         'Warehouse Manager', 'Perm', 'Alex Rivera', '— none —', 78000, 0.20,
         None, None, None, None, 1.0, 90, 'Paid', 'Yes', 'No'),
        ('P-002', '2026-07-14', 'Priya Raman', 'Cedar Health', 'RN — Nights', 'Contract',
         'Sam Chen', 'Jordan Blake', None, None, 78, 52, 36, 26, 0.7, 60, 'Started', 'No', 'No'),
        ('P-003', '2026-07-28', 'Marcus Hale', 'Fulton Manufacturing', 'CNC Machinist', 'Perm',
         'Alex Rivera', 'Jordan Blake', 64000, 0.18, None, None, None, None, 0.7, 90,
         'Started', 'No', 'No'),
    ]
    for r, row in enumerate(demo, start=2):
        (pid, dt, cand, client, role, ptype, owner, sourcer, sal, feepct,
         bill, pay, hrs, wks, split, gtee, status, paid, fell) = row
        vals = {1: pid, 2: dt, 3: cand, 4: client, 5: role, 6: ptype, 7: owner, 8: sourcer,
                9: sal, 10: feepct, 11: bill, 12: pay, 13: hrs, 14: wks,
                16: split, 19: gtee, 20: status, 21: paid, 22: fell}
        for col, v in vals.items():
            if col == 2 and isinstance(v, str):
                v = _dt.date.fromisoformat(v)
            c = pl.cell(row=r, column=col, value=v)
            c.border = BOX
            if col in (9, 11, 12):
                c.number_format = MONEY
            if col in (10, 16):
                c.number_format = PCT
            if col == 2:
                c.number_format = 'yyyy-mm-dd'

    for r in range(2, LAST):
        put = lambda col, f, fmt=MONEY: None
        # F1 gross: perm fee OR contract margin, times the house rate
        f1 = (f'=IF($A{r}="","",IF($F{r}="Perm",$I{r}*$J{r},'
              f'($K{r}-$L{r})*$M{r}*$N{r})*Settings!$B$3)')
        f2 = f'=IF($A{r}="","",$O{r}*$P{r})'                                   # owner
        f3 = f'=IF($A{r}="","",IF($H{r}="— none —",0,$O{r}*(1-$P{r})))'        # sourcer
        f4 = f'=IF($B{r}="","",$B{r}+$S{r})'                                   # guarantee ends
        f5 = (f'=IF($A{r}="","",IF($V{r}="Yes","FELL OFF",'
              f'IF(AND($B{r}<>"",TODAY()<=$W{r},$S{r}>0),"AT RISK","Cleared")))')
        f6 = f'=IF($A{r}="","",IF($V{r}="Yes",0,$O{r}))'                       # net
        for col, f, fmt in ((15, f1, MONEY), (17, f2, MONEY), (18, f3, MONEY),
                            (23, f4, 'yyyy-mm-dd'), (24, f5, None), (25, f6, MONEY)):
            c = pl.cell(row=r, column=col, value=f)
            c.fill, c.border = CALC_FILL, BOX
            if fmt:
                c.number_format = fmt

    # dropdowns, per the spec's data-validation list
    dvs = [
        (DataValidation(type='list', formula1='=Settings!$A$8:$A$9', allow_blank=True), 'F'),
        (DataValidation(type='list', formula1='=Recruiters!$A$2:$A$11', allow_blank=True), 'G'),
        (DataValidation(type='list', formula1='=Recruiters!$A$2:$A$12', allow_blank=True), 'H'),
        (DataValidation(type='list', formula1='=Settings!$B$8:$B$10', allow_blank=True), 'S'),
        (DataValidation(type='list', formula1='=Settings!$C$8:$C$11', allow_blank=True), 'T'),
        (DataValidation(type='list', formula1='"Yes,No"', allow_blank=True), 'U'),
        (DataValidation(type='list', formula1='"Yes,No"', allow_blank=True), 'V'),
    ]
    for dv, col in dvs:
        pl.add_data_validation(dv)
        dv.add(f'{col}2:{col}{LAST}')

    # ---------- Dashboard ---------------------------------------------------
    db = wb.create_sheet('Dashboard')
    db['A1'] = 'Commission Dashboard'
    db['A1'].font = Font(bold=True, size=15, color=INK)
    db['A2'] = 'Everything here is computed. Type only in Placements, Recruiters and Settings.'
    db['A2'].font = NOTE_FONT
    tiles = [
        ('Net commission — this month',
         f'=SUMIFS(Placements!$Y$2:$Y${LAST},Placements!$B$2:$B${LAST},">="&EOMONTH(TODAY(),-1)+1,'
         f'Placements!$B$2:$B${LAST},"<="&EOMONTH(TODAY(),0))'),
        ('Net commission — all time', f'=SUM(Placements!$Y$2:$Y${LAST})'),
        ('At risk (inside guarantee)',
         f'=SUMIF(Placements!$X$2:$X${LAST},"AT RISK",Placements!$Y$2:$Y${LAST})'),
        ('Placements logged', f'=COUNTA(Placements!$A$2:$A${LAST})'),
    ]
    for i, (label, formula) in enumerate(tiles):
        r = 4 + i * 3
        db.cell(row=r, column=1, value=label).font = Font(size=9, color=MUTE, bold=True)
        c = db.cell(row=r + 1, column=1, value=formula)
        c.font = Font(size=18, bold=True, color=INK)
        c.number_format = MONEY if i < 3 else '0'
    db.column_dimensions['A'].width = 34

    header(db, 17, ['Recruiter', 'Earned MTD $', 'Draw $', 'Balance $'], [22, 16, 14, 16])
    for i, r in enumerate(range(18, 21)):
        src = i + 2
        db.cell(row=r, column=1, value=f'=Recruiters!A{src}').border = BOX
        for col, ref in ((2, 'E'), (3, 'C'), (4, 'F')):
            c = db.cell(row=r, column=col, value=f'=Recruiters!{ref}{src}')
            c.number_format, c.border = MONEY, BOX

    db['A23'] = 'Deals still inside the guarantee window are marked AT RISK in Placements column X.'
    db['A23'].font = NOTE_FONT

    # ---------- Read Me -----------------------------------------------------
    rm = wb.create_sheet('Read Me')
    rm.column_dimensions['A'].width = 108
    lines = [
        ('A manual commission tracker for recruiters and staffing agencies.', 'h'),
        ('Type each placement into the Placements tab. The Dashboard rolls up who earned what, '
         "what's still pending, and which deals are still inside their guarantee window.", 'p'),
        ('', 'p'),
        ('How to use', 'h'),
        ('1. Set your people and their draws in Recruiters.', 'p'),
        ('2. Set your commission rules and dropdown options in Settings.', 'p'),
        ('3. Log every deal in Placements — only the white columns. Shaded columns compute '
         'themselves; typing over one replaces a formula and it stops updating.', 'p'),
        ('4. Read Dashboard.', 'p'),
        ('', 'p'),
        ('Two placement types', 'h'),
        ('Perm — a one-time fee, a percentage of first-year salary (columns I and J).', 'p'),
        ('Contract — ongoing margin, the spread between bill rate and pay rate '
         '(columns K to N).', 'p'),
        ('', 'p'),
        ('Two honest simplifications', 'h'),
        ('If a deal falls off, Net Commission goes to 0 rather than negative. If you were already '
         'paid out, the real clawback is money leaving your pocket — this free version does not '
         'model that.', 'p'),
        ('With no sourcer, the sourcer share stays as house margin unless you set Owner Split % '
         'to 100% on that row.', 'p'),
        ('Earned MTD credits each person with their own split — the owner gets Owner Commission, '
         'the sourcer gets Sourcer Commission — and ignores deals marked Fell Off. It is what that '
         'person earned, not the size of the deal they closed.', 'p'),
        ('', 'p'),
        ('Want it to fill itself in?', 'h'),
        ('This is the manual version — you key in every deal. The $650 Automated Dashboard pulls '
         'from your ATS or email, runs your exact commission rules, and sends an AI-written weekly '
         'commission note. I will build a working demo on your real numbers first, free, in about '
         'a day: automatedworkflowllc.com/staffing-commission-dashboard/', 'p'),
        ('', 'p'),
        ('Automated Workflow — Gainesville, FL — automatedworkflowllc.com', 'f'),
    ]
    for i, (text, kind) in enumerate(lines, start=1):
        c = rm.cell(row=i, column=1, value=text)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if kind == 'h':
            c.font = Font(bold=True, size=12, color=INK)
        elif kind == 'f':
            c.font = Font(size=9, color=MUTE)
        else:
            c.font = Font(size=10, color=INK)
        if len(text) > 90:
            rm.row_dimensions[i].height = 30

    wb.move_sheet('Read Me', offset=-4)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    return OUT


if __name__ == '__main__':
    p = build()
    print('wrote %s (%d KB)' % (p, p.stat().st_size // 1024))
