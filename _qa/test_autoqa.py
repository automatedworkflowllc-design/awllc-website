# -*- coding: utf-8 -*-
"""Regression tests for the AutoQA label-vs-value check.

Two duties, and the second matters as much as the first:
  1. the check still catches the real 7/19 Job Costing defect
  2. the check stays SILENT on the four correct formulas it originally
     misflagged — a checker nobody trusts is a checker nobody reads

Run: python _qa/test_autoqa.py
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import Workbook
from autoqa import mismatches

def sheet(headers, rows):
    wb = Workbook(); ws = wb.active
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
    for label, cell, formula in rows:
        ws.cell(row=cell, column=1, value=label)
        ws.cell(row=cell, column=4, value=formula)
    return ws

HEAD = ['Job', 'Client', 'Scope', 'Quoted $', 'Actual cost', 'Margin $', 'Margin %', 'Flag']
fails = []

def expect(name, ws, want):
    got = [m[0] for m in mismatches(ws)]
    ok = (len(got) > 0) == want
    print(('  PASS  ' if ok else '  FAIL  ') + name + ('' if ok else '  -> %s' % got))
    if not ok:
        fails.append(name)

print('AutoQA label-vs-value regression suite\n')

# --- 1. the real defect: shipped 7/19, headline read "$0 lost" above two losing jobs
expect('CATCHES the 7/19 shift: "Total quoted" summing the actual-cost column',
       sheet(HEAD, [('Total quoted', 30, '=SUM(E12:E23)')]), True)
expect('CATCHES "Total actual cost" summing the margin column',
       sheet(HEAD, [('Total actual cost', 30, '=SUM(F12:F23)')]), True)

# --- 2. the same cells, corrected
expect('SILENT on the corrected "Total quoted" -> Quoted $',
       sheet(HEAD, [('Total quoted', 30, '=SUM(D12:D23)')]), False)
expect('SILENT on the corrected "Total actual cost" -> Actual cost',
       sheet(HEAD, [('Total actual cost', 30, '=SUM(E12:E23)')]), False)

# --- 3. the four false positives from the first live run
expect('SILENT on "Actually invoiced against those bookings" -> Invoiced $',
       sheet(['Booking', 'Client', 'Date', 'Expected $', 'Inv#', 'Invoiced $', 'Delta'],
             [('Actually invoiced against those bookings', 30, '=SUM(F12:F25)')]), False)
expect('SILENT on a two-aggregate difference (delivered minus invoiced)',
       sheet(['Booking', 'Client', 'Date', 'Expected $', 'Inv#', 'Invoiced $', 'Delta'],
             [('DELIVERED BUT NOT COLLECTED', 30, '=SUM(D12:D25)-SUM(F12:F25)')]), False)
expect('SILENT on margin computed as revenue minus cost',
       sheet(['Item', 'Units', 'Price', 'Revenue', 'Handling cost', 'x', 'y'],
             [('Product margin this month', 30, '=SUM(D4:D6)-SUM(E4:E6)')]), False)
expect('SILENT on SUMIF, whose label describes the filter not the column',
       sheet(HEAD, [('LOST ON UNDER-QUOTED JOBS', 30, '=SUMIF(H12:H21,"OVER BUDGET",F12:F21)')]), False)

# --- 4. traction drift — the class that escaped three times on 7/19
from autoqa import TRACTION_DRIFT
def drift(name, text, want):
    got = bool(TRACTION_DRIFT.search(text))
    ok = got == want
    print(('  PASS  ' if ok else '  FAIL  ') + name)
    if not ok:
        fails.append(name)

print('\nTraction-drift regression suite\n')
drift('CATCHES "serves many businesses"',
      'a pipeline where one workflow serves many businesses', True)
drift('CATCHES "ten retainers is one maintained workflow"',
      'ten retainers is one maintained workflow, not ten', True)
drift('SILENT on the honest conditional "ten clients would be"',
      'ten clients would be one maintained workflow, not ten', False)
drift('SILENT on capacity phrasing "can serve many businesses"',
      'a single workflow can serve many businesses', False)
drift('SILENT on "multi-tenant by design"',
      'one workflow, multi-tenant by design', False)

# --- 5. false-proof: the QUANTITY is not the claim, the NOUN is
# "hundreds of" and "dozens of" used to match bare, with no regard for what was
# being counted. The only match site-wide on 2026-08-17 was "hundreds of Go
# modules" in the build log -- counting dependencies, implying nothing about
# clients. An engineering log will keep saying "dozens of tests" and "hundreds of
# rows"; a check that fires on those gets muted, and then it is not there on the
# day someone writes "trusted by hundreds of businesses".
from autoqa import FALSE_PROOF
def proof(name, text, want):
    got = bool(FALSE_PROOF.search(text))
    ok = got == want
    print(('  PASS  ' if ok else '  FAIL  ') + name + ('' if ok else '  -> %r' % FALSE_PROOF.findall(text)))
    if not ok:
        fails.append(name)

print('\nFalse-proof regression suite\n')
proof('CATCHES "trusted by hundreds of businesses"',
      'trusted by hundreds of businesses across Florida', True)
proof('CATCHES "dozens of clients"',
      'we serve dozens of clients across Florida', True)
proof('CATCHES "hundreds of business owners"',
      'hundreds of business owners rely on us', True)
proof('CATCHES "dozens of companies"',
      'dozens of companies use this every week', True)
proof('CATCHES "our clients"', 'our clients love it', True)
proof('CATCHES "plenty of customers"', 'plenty of customers already', True)
proof('SILENT on "hundreds of Go modules" (the 8/17 false positive)',
      'the Collector (hundreds of Go modules) and ClickHouse', False)
proof('SILENT on "dozens of tests"', 'dozens of tests now cover this', False)
proof('SILENT on "hundreds of rows"', 'hundreds of rows in the export', False)
proof('SILENT on "hundreds of pages"', 'hundreds of pages were scanned', False)
proof('SILENT on "dozens of receipts"', 'dozens of receipts in the ledger', False)
proof('SILENT on "hundreds of resources"', 'hundreds of resources in the directory', False)

# --- 6. check_deploy: silent through churn, loud when genuinely stuck
# Shipped 2026-08-17 and was wrong the same day. It called two builds "errored"
# that had not failed -- both were CANCELLED, superseded when a second push
# landed minutes after the first. Reporting routine churn as a failure is crying
# wolf, and this tree has already paid for that once: the five unread alerts from
# a monitor that fired on a schedule were the real cost, not the red badge.
#
# So these assert the SHAPE of the answer, not the state. Out of step for two
# minutes is normal and must be silent; out of step for ninety is a fact and must
# be named. The proof for this lived only in a transcript until now, which is the
# same disposable-control problem the nightly suite sweep just had fixed.
import datetime as _dt
import autoqa as _aq


class _R:
    def __init__(self, out):
        self.stdout = out


def deploy(name, head, api, want_defect):
    _aq.defects = []
    real = _aq.subprocess.run
    _aq.subprocess.run = lambda cmd, **kw: _R((head + '\n') if cmd[0] == 'git' else (api + '\n'))
    try:
        _aq.check_deploy()
    finally:
        _aq.subprocess.run = real
    got = len(_aq.defects) > 0
    ok = got == want_defect
    print(('  PASS  ' if ok else '  FAIL  ') + name +
          ('' if ok else '  -> %r' % [d['msg'] for d in _aq.defects]))
    if not ok:
        fails.append(name)


print('\ncheck_deploy regression suite\n')
_now = _dt.datetime.now(_dt.timezone.utc)
_fresh = (_now - _dt.timedelta(minutes=2)).strftime('%Y-%m-%dT%H:%M:%SZ')
_old = (_now - _dt.timedelta(minutes=90)).strftime('%Y-%m-%dT%H:%M:%SZ')
_sha, _prev = 'a' * 40, 'b' * 40

deploy('SILENT when built and matching', _sha, 'built ' + _sha + ' ' + _fresh, False)
deploy('SILENT while a build is 2 minutes old (normal churn)',
       _sha, 'building ' + _sha + ' ' + _fresh, False)
deploy('SILENT on an errored build for a SUPERSEDED commit (the 8/17 false alarm)',
       _sha, 'errored ' + _prev + ' ' + _fresh, False)
deploy('REPORTS an errored build on head after 90 minutes',
       _sha, 'errored ' + _sha + ' ' + _old, True)
deploy('REPORTS a build stuck building for 90 minutes',
       _sha, 'building ' + _sha + ' ' + _old, True)
deploy('REPORTS a live site 90 minutes behind origin/main',
       _sha, 'built ' + _prev + ' ' + _old, True)
deploy('REPORTS an unreadable API response rather than assuming healthy',
       _sha, 'nonsense', True)

print('\n%s' % ('ALL PASS' if not fails else 'FAILED: %s' % ', '.join(fails)))
sys.exit(1 if fails else 0)
