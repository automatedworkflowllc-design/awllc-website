# -*- coding: utf-8 -*-
"""AutoQA — the bounded self-checking loop for AWLLC artifacts.

Modelled on Karpathy's AutoResearch structure: an objective metric, a defined
codebase, and hard boundaries on what may change. Ours differs in one deliberate
way — it is aimed at VERIFICATION, not generation. AWLLC's constraint is demand,
not output; a loop that generates more artifacts would accelerate the thing that
is not scarce.

OBJECTIVE METRIC : total defect count (lower is better). Written to history.json
                   so regressions are detectable across runs.
CODEBASE         : awllc-website/ (demo builders + published pages)
BOUNDARIES       : READ-ONLY AND REPORT-ONLY. This script must never push, send,
                   publish, upload, or mutate a live surface. It rebuilds
                   workbooks into their own folders (idempotent) and reads. That
                   is the whole permitted blast radius. If a future edit gives it
                   write access to anything user-facing, that is a bug.

Run:  python _qa/autoqa.py           (human-readable)
      python _qa/autoqa.py --json    (machine)
Exit: 0 = no regression vs last run · 1 = regression · 2 = runner failure
"""
import io, os, re, sys, json, glob, subprocess, datetime

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HIST = os.path.join(ROOT, '_qa', 'history.json')
LIVE = 'https://automatedworkflowllc.com'

BRAND = {'211D14','FBFAF3','F4F1E8','5C5645','6E6555','E4DFD1','D8D2C2','CBC5B1',
         'FFFFFF','1E7A47','B45309','B23B3B','E6F2EC','FBEFE0','F6E7E7'}
# claims that would imply a client base AWLLC does not have
# non-capturing groups: re.findall returns tuples when a pattern has groups,
# which printed "('', '')" instead of the offending phrase
FALSE_PROOF = re.compile(r'plenty of (?:clients|customers)|many (?:clients|customers)|'
                         r'our clients|trusted by|hundreds of|dozens of', re.I)

defects = []
def defect(area, msg):
    defects.append({'area': area, 'msg': msg})

# ---------------------------------------------------------------- workbooks
def check_workbooks():
    try:
        from openpyxl import load_workbook
    except ImportError:
        defect('setup', 'openpyxl missing — workbook checks skipped'); return
    for f in sorted(glob.glob(os.path.join(ROOT, '*', '*.xlsx'))):
        rel = os.path.relpath(f, ROOT)
        try:
            wb = load_workbook(f)
        except Exception as e:
            defect('workbook', '%s unreadable: %s' % (rel, e)); continue

        # (1) conditional formatting breaks Google's xlsx import
        cf = sum(len(list(ws.conditional_formatting)) for ws in wb)
        if cf:
            defect('workbook', '%s has %d conditional-format rules (breaks Sheets import)' % (rel, cf))

        # (2) every colour must be a canonical brand token
        seen = set()
        for ws in wb:
            for row in ws.iter_rows():
                for c in row:
                    for h in (getattr(c.fill.fgColor, 'rgb', None),
                              getattr(c.font.color, 'rgb', None)):
                        if isinstance(h, str) and len(h) == 8 and h[2:].upper() != '000000':
                            seen.add(h[2:].upper())
        off = sorted(seen - BRAND)
        if off:
            defect('workbook', '%s off-brand colours: %s' % (rel, ','.join(off[:5])))

        # (3) label-vs-value — the 7/19 column-shift class
        for ws in wb:
            for label, col, head, formula in mismatches(ws):
                defect('workbook',
                       '%s [%s] "%s" sums column %s which is headed "%s"'
                       % (rel, ws.title, label[:44], col, head))


# Vocabulary of concepts a summary label and a column header can both name.
# A mismatch is only meaningful between two DIFFERENT known concepts — that is
# the exact signature of the Job Costing defect ("Total quoted" summing the
# actual-cost column). Word-boundary matching, because "Actually invoiced"
# contains "actual" and is not a defect.
CONCEPTS = {
    'quoted':    r'\bquote(d)?\b',
    'actual':    r'\bactual\b|\bactual cost\b',
    'margin':    r'\bmargin\b',
    'invoiced':  r'\binvoice(d)?\b',
    'collected': r'\bcollect(ed)?\b',
    'expense':   r'\bexpense(s)?\b',
    'revenue':   r'\brevenue\b',
    'hours':     r'\bhours?\b',
}

def concepts_in(text):
    t = (text or '').lower()
    return {name for name, pat in CONCEPTS.items() if re.search(pat, t)}

def mismatches(ws):
    """Yield (label, column, header, formula) for summary cells whose label names
    a different concept than the single column they aggregate.

    Deliberately narrow. Only plain single-column SUM/AVERAGE qualifies:
      - an expression combining two aggregates (SUM(D)-SUM(E)) computes a derived
        quantity, so its label legitimately names neither column
      - SUMIF's sum-range is a filtered subset, so the label describes the filter
    Catching four of the five real defects with zero false alarms beats catching
    five with four — a checker that cries wolf gets ignored, which is worse than
    no checker at all.
    """
    headers = {}
    for r in range(1, 16):
        filled = [c for c in ws[r] if isinstance(c.value, str) and c.value.strip()
                  and not c.value.startswith('=')]
        if len(filled) >= 4:
            for c in filled:
                headers[c.column_letter] = str(c.value)
    if not headers:
        return
    for row in ws.iter_rows():
        for c in row:
            f = c.value
            if not (isinstance(f, str) and f.startswith('=')):
                continue
            m = re.fullmatch(r'=(SUM|AVERAGE)\(([A-Z])\d+:\2\d+\)', f.strip())
            if not m:
                continue                       # derived / filtered → not checkable
            col = m.group(2)
            head = headers.get(col)
            if not head:
                continue
            label = None
            for cc in row:
                if (cc.column < c.column and isinstance(cc.value, str)
                        and cc.value.strip() and not cc.value.startswith('=')):
                    label = cc.value.strip()
            if not label:
                continue
            lab_c, head_c = concepts_in(label), concepts_in(head)
            # both must name something known, share nothing, and disagree
            if lab_c and head_c and not (lab_c & head_c):
                yield label, col, head, f

# ---------------------------------------------------------------- pages
def check_pages():
    for f in sorted(glob.glob(os.path.join(ROOT, '*', 'index.html'))
                    + glob.glob(os.path.join(ROOT, 'index.html'))):
        rel = os.path.relpath(f, ROOT)
        s = io.open(f, encoding='utf-8', errors='replace').read()
        for m in set(FALSE_PROOF.findall(s)):
            defect('page', '%s implies a client base: %r' % (rel, m))
        # every ld+json block must parse (template-clone trap)
        for blk in re.findall(r'<script type="application/ld\+json">(.*?)</script>', s, re.S):
            try:
                json.loads(blk)
            except Exception:
                defect('page', '%s has invalid ld+json' % rel)
        # a published page must not still be noindex while sitemapped
        sm = os.path.join(ROOT, 'sitemap.xml')
        if os.path.exists(sm):
            smtxt = io.open(sm, encoding='utf-8').read()
            slug = os.path.basename(os.path.dirname(f))
            if slug and slug + '/' in smtxt and 'noindex' in s:
                defect('page', '%s is in sitemap but still noindex' % rel)

# ---------------------------------------------------------------- staleness
def content_digest(path):
    """Hash the sheet content of an xlsx, ignoring docProps/.

    An xlsx is a zip, and openpyxl stamps a fresh created/modified time into
    docProps/core.xml on every write — so a plain byte compare reports every
    rebuilt file as stale even when not one cell changed. That fired on 5 of 8
    demos and would have had me overwrite good copies on a meaningless signal.
    """
    import zipfile, hashlib
    z = zipfile.ZipFile(path)
    names = [n for n in sorted(z.namelist()) if not n.startswith('docProps/')]
    return hashlib.sha256(b''.join(n.encode() + z.read(n) for n in names)).hexdigest()

def check_staleness():
    dl = os.path.join(os.path.expanduser('~'), 'Downloads')
    if not os.path.isdir(dl):
        return
    for f in sorted(glob.glob(os.path.join(ROOT, '*', '*.xlsx'))):
        b = os.path.basename(f)
        d = os.path.join(dl, b)
        if not os.path.exists(d):
            continue
        try:
            if content_digest(f) != content_digest(d):
                defect('stale', 'Downloads/%s differs from the build — sharing it sends the wrong version' % b)
        except Exception as e:
            defect('stale', 'could not compare Downloads/%s: %s' % (b, e))

# ---------------------------------------------------------------- live site
def check_live():
    sm = os.path.join(ROOT, 'sitemap.xml')
    if not os.path.exists(sm):
        return
    urls = re.findall(r'<loc>([^<]+)</loc>', io.open(sm, encoding='utf-8').read())
    for u in urls:
        try:
            out = subprocess.run(['curl', '-s', '-o', os.devnull, '-w', '%{http_code}',
                                  '-A', 'Mozilla/5.0', u],
                                 capture_output=True, text=True, timeout=25).stdout.strip()
        except Exception as e:
            defect('live', 'could not reach %s (%s)' % (u, e)); continue
        if out != '200':
            defect('live', '%s returns %s' % (u, out))

# ---------------------------------------------------------------- main
def main():
    # --fast skips the network sweep. Used by the pre-push hook, where the live
    # site is still the OLD build and so tells you nothing about what you're
    # about to ship.
    checks = [check_workbooks, check_pages, check_staleness]
    if '--fast' not in sys.argv:
        checks.append(check_live)
    for fn in checks:
        try:
            fn()
        except Exception as e:
            defect('runner', '%s crashed: %s' % (fn.__name__, e))

    score = len(defects)
    mode = 'fast' if '--fast' in sys.argv else 'full'
    hist = []
    if os.path.exists(HIST):
        try: hist = json.load(io.open(HIST, encoding='utf-8'))
        except Exception: hist = []
    # compare like with like — a fast run inspects fewer surfaces, so scoring it
    # against a full run would read as an improvement that never happened
    same = [h for h in hist if h.get('mode', 'full') == mode]
    prev = same[-1]['score'] if same else None
    stamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    hist.append({'when': stamp, 'mode': mode, 'score': score,
                 'by_area': {a: sum(1 for d in defects if d['area'] == a)
                             for a in {d['area'] for d in defects}}})
    os.makedirs(os.path.dirname(HIST), exist_ok=True)
    io.open(HIST, 'w', encoding='utf-8').write(json.dumps(hist[-60:], indent=1))

    if '--json' in sys.argv:
        print(json.dumps({'score': score, 'prev': prev, 'defects': defects}, indent=1))
    else:
        print('AutoQA %s  —  defects: %s%s' % (
            stamp, score,
            '' if prev is None else '  (previous: %s)' % prev))
        for d in defects:
            print('  [%-8s] %s' % (d['area'], d['msg']))
        if not defects:
            print('  clean')
        if prev is not None and score > prev:
            print('\n  ** REGRESSION: %d new defect(s) since last run **' % (score - prev))

    if prev is not None and score > prev:
        return 1
    return 0

if __name__ == '__main__':
    try:
        sys.exit(main())
    except Exception as e:
        print('runner failure:', e); sys.exit(2)
