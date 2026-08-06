# -*- coding: utf-8 -*-
"""Build the SELLABLE template variant of the Money Leak Finder.

Differs from the demo (build_money_leak.py) in exactly the ways a paid product demands:
  * EVERY flag/bucket/KPI is a LIVE FORMULA, not a Python-computed static — a buyer who
    pastes their own rows gets a report that actually updates. (Selling "self-updating"
    with baked statics would be the label-vs-value defect as a product.)
  * Ranges extend to row 300 so added rows are included without editing formulas.
  * HOW TO USE tab: honest instructions incl. FILTER() needing Excel 365 / Google Sheets,
    and that sample-row colors are illustrative (no conditional formatting — Sheets drops
    it on xlsx import, so flags are formula text, not color magic).
  * Sample data kept (buyers orient faster with an example) and labeled as invented.
QA rules: self-reconciliation printed below; every summary formula aggregates the column
its label names; colors derived from the same logic the formulas use.
"""
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import sys as _sys, os as _os
_sys.path.insert(0, _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "..", "_brand"))
from awllc_brand import (INK, GREEN, AMBER, RED, MUTE, CARD, WORDMARK,
                         GREENBG, REDBG, AMBERBG, LINE, PAPER_TINT)

AS_OF = dt.date(2026, 7, 28)   # sample as-of; buyer replaces with =TODAY() per instructions
MAXR = 300

HEADER_FILL = PatternFill("solid", fgColor=INK)
CARD_FILL   = PatternFill("solid", fgColor=CARD)
GREEN_FILL  = PatternFill("solid", fgColor=GREENBG)
RED_FILL    = PatternFill("solid", fgColor=REDBG)
AMBER_FILL  = PatternFill("solid", fgColor=AMBERBG)
thin = Side(style="thin", color=LINE)

wb = Workbook()

def band(ws, rng, text, size=13):
    ws.merge_cells(rng)
    top = rng.split(":")[0]
    ws[top] = text
    ws[top].font = Font(bold=True, size=size, color="FFFFFF")
    ws[top].alignment = Alignment(horizontal="left", vertical="center")
    from openpyxl.utils.cell import range_boundaries
    minc, minr, maxc, maxr = range_boundaries(rng)
    for r in range(minr, maxr + 1):
        for c in range(minc, maxc + 1):
            ws.cell(r, c).fill = HEADER_FILL
    ws.row_dimensions[minr].height = 28

def headers_row(ws, r, cols):
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
    ws.row_dimensions[r].height = 20

def wordmark(ws, row):
    ws.cell(row, 1, "Template by Automated Workflow · automatedworkflowllc.com")
    ws.cell(row, 1).font = Font(italic=True, size=9, color=WORDMARK)

# ================= sample data (same invented scenario as the demo) =================
JOBS = [
    ("J-101", dt.date(2026, 6, 22), "Hargrove Property Grp", "Irrigation main repair",        740.00),
    ("J-102", dt.date(2026, 6, 25), "Bell & Sons HOA",       "Storm cleanup, common area",    1280.00),
    ("J-103", dt.date(2026, 6, 28), "Rivertown Storage",     "Gate motor replacement",         965.00),
    ("J-104", dt.date(2026, 7, 1),  "Hargrove Property Grp", "Emergency leak call-out",        410.00),
    ("J-105", dt.date(2026, 7, 2),  "Coastal Dental",        "Parking lot pressure wash",      380.00),
    ("J-106", dt.date(2026, 7, 6),  "Bell & Sons HOA",       "Fence section rebuild",         1620.00),
    ("J-107", dt.date(2026, 7, 8),  "Marlin Self-Serve",     "Bay door track repair",          540.00),
    ("J-108", dt.date(2026, 7, 9),  "Rivertown Storage",     "Unit door replacements (x3)",   1150.00),
    ("J-109", dt.date(2026, 7, 12), "Coastal Dental",        "After-hours lighting fix",       295.00),
    ("J-110", dt.date(2026, 7, 14), "Northgate Church",      "Grounds cleanup + haul-away",    860.00),
    ("J-111", dt.date(2026, 7, 16), "Hargrove Property Grp", "Weekend emergency call-out",     520.00),
    ("J-112", dt.date(2026, 7, 18), "Marlin Self-Serve",     "Camera pole reset",              310.00),
    ("J-113", dt.date(2026, 7, 21), "Bell & Sons HOA",       "Playground mulch install",       990.00),
    ("J-114", dt.date(2026, 7, 24), "Northgate Church",      "Gutter clear, full building",    450.00),
]
INVOICES = [
    ("INV-2201", "J-101", dt.date(2026, 6, 24), 740.00,  dt.date(2026, 7, 2)),
    ("INV-2202", "J-102", dt.date(2026, 6, 27), 1280.00, dt.date(2026, 7, 18)),
    ("INV-2203", "J-103", dt.date(2026, 6, 30), 965.00,  None),
    ("INV-2204", "J-105", dt.date(2026, 7, 4),  380.00,  dt.date(2026, 7, 11)),
    ("INV-2205", "J-106", dt.date(2026, 7, 8),  1620.00, None),
    ("INV-2206", "J-107", dt.date(2026, 7, 10), 540.00,  dt.date(2026, 7, 22)),
    ("INV-2207", "J-108", dt.date(2026, 7, 11), 1150.00, None),
    ("INV-2208", "J-110", dt.date(2026, 7, 15), 860.00,  dt.date(2026, 7, 24)),
    ("INV-2209", "J-113", dt.date(2026, 7, 22), 990.00,  None),
    ("INV-2210", "J-104", dt.date(2026, 5, 12), 410.00,  None),
    ("INV-2211", "J-109", dt.date(2026, 4, 20), 295.00,  None),
]

# ================= Tab 1: HOW TO USE =================
ht = wb.active
ht.title = "How To Use"
band(ht, "A1:B1", "MONEY LEAK FINDER · how to use this template", size=14)
LINES = [
    ("", ""),
    ("What this does", "Puts work you COMPLETED next to work you INVOICED and flags only the mismatches: "
     "jobs never billed, invoices unpaid, and how long they've been aging."),
    ("", ""),
    ("Step 1", "Work Log tab: replace the sample rows with your completed jobs (one row per job). "
     "Keep the Job ID column — it's how the sheets link. Sample data is invented; delete it freely."),
    ("Step 2", "Invoices tab: paste your invoices (one row per invoice). Put the matching Job ID in "
     "column B. Leave 'Date paid' empty until an invoice is paid."),
    ("Step 3", "Leak Report tab: set the As-of date cell (B3). Type a date, or enter =TODAY() to make "
     "the report recalculate itself every time you open the file."),
    ("Step 4", "Read only the flags: NEVER BILLED rows on the Work Log, UNPAID rows and their aging "
     "bucket on the Invoices tab, and the three totals on the Leak Report."),
    ("", ""),
    ("Capacity", "Formulas cover rows up to 300 on both data tabs. Past that, extend the ranges "
     "(or ask us — contact below)."),
    ("Honest notes", "1) The 'never billed' list on the Leak Report uses FILTER(), which needs "
     "Google Sheets or Excel 365. On older Excel that one cell shows #NAME? — everything else still "
     "works; use the NEVER BILLED flag column on the Work Log instead. "
     "2) Colored fills on the sample rows are illustrative only; the flags themselves are formulas "
     "and always update. 3) All sample names and figures are invented."),
    ("", ""),
    ("Support", "Stuck, or want this wired to your real exports and refreshing automatically? "
     "colin@automatedworkflowllc.com — a working version from your own files inside a day, free to try."),
]
r = 2
for label, text in LINES:
    ht.cell(r, 1, label).font = Font(bold=True, color=INK)
    c = ht.cell(r, 2, text)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if text:
        ht.row_dimensions[r].height = max(15, 13 * (1 + len(text) // 95))
    r += 1
wordmark(ht, r + 1)
ht.column_dimensions["A"].width = 14
ht.column_dimensions["B"].width = 100

# ================= Tab 2: WORK LOG (live flag column) =================
ws = wb.create_sheet("Work Log")
band(ws, "A1:F1", "WORK LOG · what got done   (sample rows — replace with yours)")
ws["A2"] = "The Billed? column is a formula — it flags any job with no matching invoice. Don't overwrite column F."
ws["A2"].font = Font(italic=True, size=10, color=MUTE)
headers_row(ws, 3, ["Job ID", "Date completed", "Customer", "Description", "Amount $", "Billed?"])
for i, j in enumerate(JOBS):
    r = 4 + i
    ws.cell(r, 1, j[0]); ws.cell(r, 2, j[1].isoformat()); ws.cell(r, 3, j[2])
    ws.cell(r, 4, j[3]); ws.cell(r, 5, j[4]).number_format = '#,##0.00'
    billed = any(inv[1] == j[0] for inv in INVOICES)
    if not billed:
        for c in range(1, 6):
            ws.cell(r, c).fill = RED_FILL
    print(f"  WORKLOG {j[0]}: billed={'yes' if billed else 'NO (red sample row)'}")
# live formula for ALL rows 4..MAXR (works for buyer-added rows too)
for r in range(4, MAXR + 1):
    ws.cell(r, 6, f'=IF(A{r}="","",IF(COUNTIF(Invoices!B$4:B${MAXR},A{r})=0,"NEVER BILLED",""))')
    ws.cell(r, 6).font = Font(bold=True, color=RED)
ws.cell(2, 5, "").border = Border(bottom=thin)
tr = MAXR + 2
ws.cell(tr, 4, "TOTAL COMPLETED $").font = Font(bold=True)
ws.cell(tr, 5, f"=SUM(E4:E{MAXR})").number_format = '#,##0.00'
ws.cell(tr, 5).font = Font(bold=True)
wordmark(ws, tr + 2)
for col, w in zip("ABCDEF", [10, 15, 22, 30, 12, 14]):
    ws.column_dimensions[col].width = w

# ================= Tab 3: INVOICES (live status/days/bucket) =================
inv = wb.create_sheet("Invoices")
band(inv, "A1:H1", "INVOICES · what got billed   (sample rows — replace with yours)")
inv["A2"] = "Status, Days out and Bucket are formulas — leave columns F, G, H alone. Aging counts from the As-of date on the Leak Report."
inv["A2"].font = Font(italic=True, size=10, color=MUTE)
headers_row(inv, 3, ["Invoice #", "Job ID", "Date sent", "Amount $", "Date paid", "Status", "Days out", "Bucket"])
for i, v in enumerate(INVOICES):
    r = 4 + i
    inv.cell(r, 1, v[0]); inv.cell(r, 2, v[1])
    inv.cell(r, 3, v[2]); inv.cell(r, 3).number_format = 'yyyy-mm-dd'
    inv.cell(r, 4, v[3]).number_format = '#,##0.00'
    if v[4]:
        inv.cell(r, 5, v[4]).number_format = 'yyyy-mm-dd'
    paid = v[4] is not None
    days = (AS_OF - v[2]).days
    bucket = "PAID" if paid else ("0-30" if days <= 30 else "31-60" if days <= 60 else "61-90" if days <= 90 else "90+")
    fill = GREEN_FILL if paid else (RED_FILL if days > 60 else AMBER_FILL if days > 30 else PatternFill("solid", fgColor=PAPER_TINT))
    for c in range(1, 9):
        inv.cell(r, c).fill = fill
    print(f"  INVOICE {v[0]}: paid={paid}, days={days if not paid else '-'}, bucket={bucket} -> fill={'green' if paid else 'red' if days>60 else 'amber' if days>30 else 'neutral'}")
for r in range(4, MAXR + 1):
    inv.cell(r, 6, f'=IF(A{r}="","",IF(E{r}="","UNPAID","PAID"))')
    inv.cell(r, 7, f"=IF(OR(A{r}=\"\",F{r}=\"PAID\"),\"\",'Leak Report'!$B$3-C{r})")
    inv.cell(r, 8, f'=IF(G{r}="","",IF(G{r}<=30,"0-30",IF(G{r}<=60,"31-60",IF(G{r}<=90,"61-90","90+"))))')
    inv.cell(r, 6).font = Font(bold=True)
tr = MAXR + 2
inv.cell(tr, 3, "TOTAL INVOICED $").font = Font(bold=True)
inv.cell(tr, 4, f"=SUM(D4:D{MAXR})").number_format = '#,##0.00'
inv.cell(tr, 4).font = Font(bold=True)
wordmark(inv, tr + 2)
for col, w in zip("ABCDEFGH", [12, 10, 13, 12, 13, 10, 10, 9]):
    inv.column_dimensions[col].width = w

# ================= Tab 4: LEAK REPORT (all live) =================
lr = wb.create_sheet("Leak Report")
band(lr, "A1:F1", "LEAK REPORT · only the mismatches")
lr.cell(3, 1, "As-of date →").font = Font(bold=True, color=MUTE)
lr.cell(3, 2, AS_OF).number_format = 'yyyy-mm-dd'
lr.cell(3, 2).font = Font(bold=True)
lr.cell(3, 3, "type a date, or enter =TODAY() to self-update").font = Font(italic=True, size=9, color=MUTE)

WL = f"'Work Log'"
kpis = [
    (5, "COMPLETED, NEVER INVOICED $",
     f"=SUMIF({WL}!F4:F{MAXR},\"NEVER BILLED\",{WL}!E4:E{MAXR})", RED, RED_FILL),
    (6, "INVOICED, STILL UNPAID $",
     f'=SUMIF(Invoices!F4:F{MAXR},"UNPAID",Invoices!D4:D{MAXR})', AMBER, AMBER_FILL),
    (7, "UNPAID & PAST 60 DAYS $",
     f'=SUMIFS(Invoices!D4:D{MAXR},Invoices!F4:F{MAXR},"UNPAID",Invoices!G4:G{MAXR},">60")', RED, RED_FILL),
]
for row, label, formula, color, fill in kpis:
    lr.cell(row, 1, label).font = Font(bold=True, size=10, color=MUTE)
    cell = lr.cell(row, 3, formula)
    cell.font = Font(bold=True, size=16, color=color)
    cell.number_format = '"$"#,##0.00'
    for c in range(1, 5):
        lr.cell(row, c).fill = fill if row != 6 else AMBER_FILL

lr.cell(9, 1, "JOBS COMPLETED BUT NEVER INVOICED (needs Google Sheets or Excel 365 — see How To Use)").font = Font(bold=True, color=RED)
lr.cell(10, 1, f"=IFERROR(FILTER({WL}!A4:E{MAXR},{WL}!F4:F{MAXR}=\"NEVER BILLED\"),\"— none —\")")
lr.cell(16, 1, "UNPAID INVOICES, OLDEST FIRST — sort/filter the Invoices tab by the Bucket column").font = Font(bold=True, color=AMBER)
lr.cell(18, 1, "Every figure on this tab recalculates from your data. Sample data is invented.").font = Font(italic=True, size=10, color=MUTE)
wordmark(lr, 20)
for col, w in zip("ABCDEF", [30, 16, 18, 12, 12, 12]):
    lr.column_dimensions[col].width = w

# ================= self-reconciliation =================
print("\n=== SELF-RECONCILIATION (Python truth vs what the live formulas will compute) ===")
nb = [j for j in JOBS if not any(v[1] == j[0] for v in INVOICES)]
unpaid = [v for v in INVOICES if v[4] is None]
over60 = [v for v in unpaid if (AS_OF - v[2]).days > 60]
print(f"  never billed: {[j[0] for j in nb]} -> ${sum(j[4] for j in nb):,.2f}  (KPI 1 SUMIF over WorkLog F/E)")
print(f"  unpaid:       {[v[0] for v in unpaid]} -> ${sum(v[3] for v in unpaid):,.2f}  (KPI 2 SUMIF over Invoices F/D)")
print(f"  unpaid >60d:  {[v[0] for v in over60]} -> ${sum(v[3] for v in over60):,.2f}  (KPI 3 SUMIFS w/ G>60)")
assert len(nb) == 3 and abs(sum(j[4] for j in nb) - 1280.00) < 0.01
assert len(unpaid) == 6 and abs(sum(v[3] for v in unpaid) - 5430.00) < 0.01
assert len(over60) == 2 and abs(sum(v[3] for v in over60) - 705.00) < 0.01

out = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "money-leak-finder-TEMPLATE.xlsx")
wb.save(out)
print(f"\nsaved {out} ({_os.path.getsize(out)} bytes)")
