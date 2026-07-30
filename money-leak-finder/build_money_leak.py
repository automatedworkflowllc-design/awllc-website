# -*- coding: utf-8 -*-
"""Build "Money Leak Finder — SAMPLE (Automated Workflow LLC)" as xlsx for Drive conversion.

THE product the money-leak outreach hook promises: put two lists side by side —
work completed vs. work invoiced — and flag ONLY the mismatches:
  A. NEVER BILLED   — completed jobs with no invoice          (money earned, never asked for)
  B. UNPAID / AGING — invoices with no payment, aged 0-30/31-60/61-90/90+
Scenario: "Cedar Field Services" — an INVENTED small field-services company.
Every figure is sample data and the workbook says so on every tab.

QA rules honored (from awllc-technical-lessons, each born from a real shipped bug):
  * every summary formula aggregates the column its own label names (label-vs-value)
  * static colors are DERIVED in Python from the same logic the sheet formulas use,
    and every derivation is PRINTED below (input -> computed tier)
  * dashboard self-reconciles: Python totals printed for cross-check vs formulas
  * aging uses a FIXED as-of date (no TODAY()) so colors can never drift from math
  * no conditional formatting (Sheets drops it on xlsx import)
"""
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import sys as _sys, os as _os
_sys.path.insert(0, _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "..", "_brand"))
from awllc_brand import (INK, GREEN, AMBER, RED, MUTE, CARD, BANDBG, WORDMARK,
                         GREENBG, REDBG, AMBERBG, LINE, PAPER_TINT)

AS_OF = dt.date(2026, 7, 28)   # fixed report date — deterministic aging

HEADER_FILL = PatternFill("solid", fgColor=INK)
CARD_FILL   = PatternFill("solid", fgColor=CARD)
GREEN_FILL  = PatternFill("solid", fgColor=GREENBG)
RED_FILL    = PatternFill("solid", fgColor=REDBG)
AMBER_FILL  = PatternFill("solid", fgColor=AMBERBG)
thin = Side(style="thin", color=LINE)
bottom_border = Border(bottom=thin)

wb = Workbook()

def band(ws, rng, text, size=14):
    ws.merge_cells(rng)
    top = rng.split(":")[0]
    ws[top] = text
    ws[top].font = Font(bold=True, size=size, color="FFFFFF")
    ws[top].alignment = Alignment(horizontal="left", vertical="center")
    from openpyxl.utils.cell import range_boundaries
    minc, minr, maxc, maxr = range_boundaries(rng)
    for r in range(minr, maxr + 1):
        for c in range(minc, maxc + 1):
            ws.cell(r, c).fill = HEADER_FILL
    ws.row_dimensions[minr].height = 30

def note(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(italic=True, size=10, color=MUTE)

def headers_row(ws, r, cols):
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
    ws.row_dimensions[r].height = 22

def wordmark(ws, row):
    ws.cell(row, 1, "Built by Automated Workflow LLC  ·  automatedworkflowllc.com  ·  sample data throughout")
    ws.cell(row, 1).font = Font(italic=True, size=9, color=WORDMARK)

# ================= seed data (INVENTED) =================
# (job_id, date_done, customer, description, amount)
JOBS = [
    ("J-101", dt.date(2026, 6, 22), "Hargrove Property Grp", "Irrigation main repair",        740.00),
    ("J-102", dt.date(2026, 6, 25), "Bell & Sons HOA",       "Storm cleanup, common area",    1280.00),
    ("J-103", dt.date(2026, 6, 28), "Rivertown Storage",     "Gate motor replacement",         965.00),
    ("J-104", dt.date(2026, 7, 1),  "Hargrove Property Grp", "Emergency leak call-out",        410.00),
    ("J-105", dt.date(2026, 7, 2),  "Coastal Dental",        "Parking lot pressure wash",      380.00),
    ("J-106", dt.date(2026, 7, 6),  "Bell & Sons HOA",       "Fence section rebuild",         1620.00),
    ("J-107", dt.date(2026, 7, 8),  "Marlin Self-Serve",     "Bay door track repair",          540.00),
    ("J-108", dt.date(2026, 7, 9),  "Rivertown Storage",     "Unit door replacements (x3)",   1150.00),
    ("J-109", dt.date(2026, 7, 12), "Coastal Dental",        "After-hours lighting fix",       295.00),
    ("J-110", dt.date(2026, 7, 14), "Northgate Church",      "Grounds cleanup + haul-away",    860.00),
    ("J-111", dt.date(2026, 7, 16), "Hargrove Property Grp", "Weekend emergency call-out",     520.00),
    ("J-112", dt.date(2026, 7, 18), "Marlin Self-Serve",     "Camera pole reset",              310.00),
    ("J-113", dt.date(2026, 7, 21), "Bell & Sons HOA",       "Playground mulch install",       990.00),
    ("J-114", dt.date(2026, 7, 24), "Northgate Church",      "Gutter clear, full building",    450.00),
]
# invoices: (inv_id, job_id, date_sent, amount, date_paid or None)
INVOICES = [
    ("INV-2201", "J-101", dt.date(2026, 6, 24), 740.00,  dt.date(2026, 7, 2)),
    ("INV-2202", "J-102", dt.date(2026, 6, 27), 1280.00, dt.date(2026, 7, 18)),
    ("INV-2203", "J-103", dt.date(2026, 6, 30), 965.00,  None),                 # 28d — current-ish
    ("INV-2204", "J-105", dt.date(2026, 7, 4),  380.00,  dt.date(2026, 7, 11)),
    ("INV-2205", "J-106", dt.date(2026, 7, 8),  1620.00, None),                 # 20d
    ("INV-2206", "J-107", dt.date(2026, 7, 10), 540.00,  dt.date(2026, 7, 22)),
    ("INV-2207", "J-108", dt.date(2026, 7, 11), 1150.00, None),                 # 17d
    ("INV-2208", "J-110", dt.date(2026, 7, 15), 860.00,  dt.date(2026, 7, 24)),
    ("INV-2209", "J-113", dt.date(2026, 7, 22), 990.00,  None),                 # 6d
    ("INV-2210", "J-104", dt.date(2026, 5, 12), 410.00,  None),                 # backdated: 77d  61-90 bucket
    ("INV-2211", "J-109", dt.date(2026, 4, 20), 295.00,  None),                 # backdated: 99d  90+ bucket
]
# NOTE: INV-2210 / INV-2211 are deliberately backdated relative to their job dates to
# model re-issued/carried-over invoices — the aging story needs 61-90 and 90+ examples.
inv_by_job = {i[1] for i in INVOICES}
NEVER_BILLED = [j for j in JOBS if j[0] not in inv_by_job]      # J-111, J-112, J-114 expected

def bucket(days):
    return "0-30" if days <= 30 else "31-60" if days <= 60 else "61-90" if days <= 90 else "90+"

# ================= Tab 1: WORK LOG =================
ws = wb.active
ws.title = "Work Log"
band(ws, "A1:E1", "WORK LOG  ·  what got DONE   —   SAMPLE DATA (every figure invented)")
note(ws, "A2", "One row per completed job. In the live version this comes from your scheduler/job system.")
headers_row(ws, 3, ["Job ID", "Date completed", "Customer", "Description", "Amount $"])
r = 4
for j in JOBS:
    ws.cell(r, 1, j[0]); ws.cell(r, 2, j[1].isoformat()); ws.cell(r, 3, j[2])
    ws.cell(r, 4, j[3]); ws.cell(r, 5, j[4]).number_format = '#,##0.00'
    ws.cell(r, 5).border = bottom_border
    # derived flag: billed or not — same COUNTIF logic the Leak Report uses
    billed = j[0] in inv_by_job
    if not billed:
        for c in range(1, 6):
            ws.cell(r, c).fill = RED_FILL
    print(f"  WORKLOG {j[0]}: billed={'yes' if billed else 'NO -> red row'}")
    r += 1
ws.cell(r + 1, 1, f"TOTAL WORK COMPLETED $"); ws.cell(r + 1, 1).font = Font(bold=True)
ws.cell(r + 1, 5, f"=SUM(E4:E{r-1})").number_format = '#,##0.00'
ws.cell(r + 1, 5).font = Font(bold=True)
wordmark(ws, r + 3)
for col, w in zip("ABCDE", [10, 15, 22, 30, 12]):
    ws.column_dimensions[col].width = w

# ================= Tab 2: INVOICES =================
inv = wb.create_sheet("Invoices")
band(inv, "A1:F1", "INVOICES  ·  what got BILLED   —   SAMPLE DATA (every figure invented)")
note(inv, "A2", "One row per invoice sent. In the live version this is your invoice/accounting export.")
headers_row(inv, 3, ["Invoice #", "Job ID", "Date sent", "Amount $", "Date paid", "Status"])
r = 4
for i in INVOICES:
    inv.cell(r, 1, i[0]); inv.cell(r, 2, i[1]); inv.cell(r, 3, i[2].isoformat())
    inv.cell(r, 4, i[3]).number_format = '#,##0.00'
    inv.cell(r, 5, i[4].isoformat() if i[4] else "")
    # status derived from date_paid — same logic Leak Report uses
    paid = i[4] is not None
    inv.cell(r, 6, "PAID" if paid else "UNPAID")
    inv.cell(r, 6).font = Font(bold=True, color=(GREEN if paid else RED))
    inv.cell(r, 6).fill = GREEN_FILL if paid else RED_FILL
    days = (AS_OF - i[2]).days
    print(f"  INVOICE {i[0]} ({i[1]}): paid={paid}, days_out={days}, bucket={bucket(days) if not paid else '-'}")
    r += 1
inv.cell(r + 1, 1, "TOTAL INVOICED $"); inv.cell(r + 1, 1).font = Font(bold=True)
inv.cell(r + 1, 4, f"=SUM(D4:D{r-1})").number_format = '#,##0.00'
inv.cell(r + 1, 4).font = Font(bold=True)
wordmark(inv, r + 3)
for col, w in zip("ABCDEF", [12, 10, 14, 12, 14, 10]):
    inv.column_dimensions[col].width = w

# ================= Tab 3: LEAK REPORT =================
lr = wb.create_sheet("Leak Report")
band(lr, "A1:F1", f"LEAK REPORT  ·  as of {AS_OF.isoformat()}   —   SAMPLE DATA (every figure invented)")
note(lr, "A2", "Only the mismatches. Live version recalculates itself; this sample is frozen at the as-of date above.")

# --- KPI cards (formulas reference the sections below so labels can't drift) ---
py_never_billed_total = sum(j[4] for j in NEVER_BILLED)
unpaid = [(i, (AS_OF - i[2]).days) for i in INVOICES if i[4] is None]
py_unpaid_total = sum(i[0][3] for i in unpaid)
py_over60_total = sum(i[0][3] for i in unpaid if i[1] > 60)

n_nb = len(NEVER_BILLED)
nb_first, nb_last = 8, 8 + n_nb - 1                        # section A data rows
up_first = nb_last + 5                                     # section B header offset computed below

lr.cell(3, 1, "COMPLETED, NEVER INVOICED").font = Font(bold=True, size=10, color=MUTE)
lr.cell(4, 1, f"=SUM(E{nb_first}:E{nb_last})").font = Font(bold=True, size=18, color=RED)
lr.cell(4, 1).number_format = '"$"#,##0.00'
lr.cell(3, 3, "INVOICED, STILL UNPAID").font = Font(bold=True, size=10, color=MUTE)
lr.cell(3, 5, "UNPAID & OLDER THAN 60 DAYS").font = Font(bold=True, size=10, color=MUTE)
for cell in ("A3", "A4", "C3", "C4", "E3", "E4"):
    lr[cell].fill = CARD_FILL

# --- Section A: never billed ---
lr.cell(6, 1, "A · COMPLETED BUT NEVER INVOICED  (from Work Log with no matching invoice)").font = Font(bold=True, color=RED)
headers_row(lr, 7, ["Job ID", "Date completed", "Customer", "Description", "Amount $"])
r = nb_first
for j in NEVER_BILLED:
    lr.cell(r, 1, j[0]); lr.cell(r, 2, j[1].isoformat()); lr.cell(r, 3, j[2])
    lr.cell(r, 4, j[3]); lr.cell(r, 5, j[4]).number_format = '#,##0.00'
    for c in range(1, 6):
        lr.cell(r, c).fill = RED_FILL
    print(f"  LEAK-A {j[0]}: never billed -> ${j[4]:,.2f}")
    r += 1

# --- Section B: unpaid + aging ---
sb = r + 2
lr.cell(sb, 1, "B · INVOICED BUT UNPAID  (aged from date sent to the as-of date)").font = Font(bold=True, color=AMBER)
headers_row(lr, sb + 1, ["Invoice #", "Customer", "Date sent", "Amount $", "Days out", "Bucket"])
r = sb + 2
up_first_row = r
cust_by_job = {j[0]: j[2] for j in JOBS}
for (i, days) in sorted(unpaid, key=lambda x: -x[1]):
    b = bucket(days)
    fill = RED_FILL if days > 60 else AMBER_FILL if days > 30 else PatternFill("solid", fgColor=PAPER_TINT)
    color = RED if days > 60 else AMBER if days > 30 else MUTE
    lr.cell(r, 1, i[0]); lr.cell(r, 2, cust_by_job[i[1]]); lr.cell(r, 3, i[2].isoformat())
    lr.cell(r, 4, i[3]).number_format = '#,##0.00'
    lr.cell(r, 5, days); lr.cell(r, 6, b)
    lr.cell(r, 6).font = Font(bold=True, color=color)
    for c in range(1, 7):
        lr.cell(r, c).fill = fill
    print(f"  LEAK-B {i[0]}: {days}d -> bucket {b} -> {'red' if days>60 else 'amber' if days>30 else 'neutral'}")
    r += 1
up_last_row = r - 1

# KPI formulas now that section B rows are known
lr.cell(4, 3, f"=SUM(D{up_first_row}:D{up_last_row})").font = Font(bold=True, size=18, color=AMBER)
lr.cell(4, 3).number_format = '"$"#,##0.00'
lr.cell(4, 5, f'=SUMIF(E{up_first_row}:E{up_last_row},">60",D{up_first_row}:D{up_last_row})').font = Font(bold=True, size=18, color=RED)
lr.cell(4, 5).number_format = '"$"#,##0.00'

# --- footer ---
fr = r + 2
lr.cell(fr, 1, "How the live version differs: it reads your real systems, recalculates every day, and emails you only when something new shows up.")
lr.cell(fr, 1).font = Font(italic=True, size=10, color=MUTE)
wordmark(lr, fr + 2)
for col, w in zip("ABCDEF", [12, 20, 14, 12, 10, 10]):
    lr.column_dimensions[col].width = w

# ================= self-reconciliation printout =================
print("\n=== SELF-RECONCILIATION (Python truth vs what the sheet formulas will compute) ===")
print(f"  never-billed jobs: {[j[0] for j in NEVER_BILLED]}  total ${py_never_billed_total:,.2f}")
print(f"  unpaid invoices:   {[i[0][0] for i in unpaid]}  total ${py_unpaid_total:,.2f}")
print(f"  unpaid >60 days:   {[i[0][0] for i in unpaid if i[1] > 60]}  total ${py_over60_total:,.2f}")
print(f"  KPI A (never invoiced) formula sums LeakReport E{nb_first}:E{nb_last} -> expect {py_never_billed_total:,.2f}")
print(f"  KPI B (unpaid) formula sums D{up_first_row}:D{up_last_row} -> expect {py_unpaid_total:,.2f}")
print(f"  KPI C (>60) SUMIF on E{up_first_row}:E{up_last_row} -> expect {py_over60_total:,.2f}")
assert abs(py_never_billed_total - 1280.00) < 0.01, "expected J-111+J-112+J-114 = 520+310+450"
assert len(unpaid) == 6 and len(NEVER_BILLED) == 3

out = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "money-leak-finder-sample.xlsx")
wb.save(out)
print(f"\nsaved {out} ({_os.path.getsize(out)} bytes)")
